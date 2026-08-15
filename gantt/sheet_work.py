"""The Tasks and Sub-Tasks input tabs, including the derived columns."""

from openpyxl.formatting.rule import Rule
from openpyxl.styles import Border, Font, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

from . import demo, layout as L, names as N, styles as S


def _list_validation(ws, name: str, col: int, first_row: int, last_row: int,
                     allow_blank: bool = True) -> None:
    dv = DataValidation(type="list", formula1=f"={name}", allow_blank=allow_blank,
                        showDropDown=False)
    dv.error = f"Pick a value from the {name} list."
    dv.errorTitle = "Not a known value"
    ws.add_data_validation(dv)
    dv.add(f"{L.col(col)}{first_row}:{L.col(col)}{last_row}")


def _check_formatting(ws, col: int, first_row: int, last_row: int) -> None:
    rng = f"{L.col(col)}{first_row}:{L.col(col)}{last_row}"
    ws.conditional_formatting.add(rng, Rule(
        type="containsText", operator="containsText", text="⚠",
        dxf=DifferentialStyle(fill=S.CF_WARN),
        formula=[f'NOT(ISERROR(SEARCH("⚠",{L.col(col)}{first_row})))']))
    ws.conditional_formatting.add(rng, Rule(
        type="containsText", operator="containsText", text="ok",
        dxf=DifferentialStyle(fill=S.CF_OK),
        formula=[f'NOT(ISERROR(SEARCH("ok",{L.col(col)}{first_row})))']))


# --- Tasks -----------------------------------------------------------------

def build_tasks(ws) -> None:
    S.header_row(ws, 1, 1, [L.TASK_HEADERS[c] for c in sorted(L.TASK_HEADERS)])
    for c in L.TASK_DERIVED_COLS:
        cell = ws.cell(row=1, column=c)
        cell.fill = S.FILL_OUTPUT_HDR

    for i, task in enumerate(demo.TASKS):
        r = L.TASK_FIRST_ROW + i
        tid, name, category, priority, complexity, equipment, assignee, start_ww, _ = task
        ws.cell(row=r, column=L.T_ID, value=tid)
        ws.cell(row=r, column=L.T_NAME, value=name)
        ws.cell(row=r, column=L.T_CATEGORY, value=category)
        ws.cell(row=r, column=L.T_PRIORITY, value=priority)
        ws.cell(row=r, column=L.T_COMPLEXITY, value=complexity)
        ws.cell(row=r, column=L.T_EQUIPMENT, value=equipment)
        ws.cell(row=r, column=L.T_DEF_ASSIGNEE, value=assignee)
        ws.cell(row=r, column=L.T_START_WW, value=start_ww)

    for r in range(L.TASK_FIRST_ROW, N.LAST_TASK_ROW + 1):
        # Letters are looked up rather than written in: a column inserted
        # anywhere to the left silently rewrites what every hardcoded one means.
        c_id = L.col(L.T_ID)
        c_start = L.col(L.T_START_WW)
        c_subs = L.col(L.T_N_SUBS)
        c_remaining = L.col(L.T_REMAINING)
        c_status = L.col(L.T_STATUS)
        tw_row = L.CW_TASKWEEK_FIRST + (r - L.TASK_FIRST_ROW)
        span = (f"{L.sheet_ref(L.CALC_WEEK)}!"
                f"${L.col(L.CW_FIRST_WEEK_COL)}${tw_row}:${L.col(N.CW_LAST_WEEK_COL)}${tw_row}")
        effort = f"SUMIF(SubParent,${c_id}{r},SubEffort)"

        ws.cell(row=r, column=L.T_N_SUBS,
                value=f'=IF(${c_id}{r}="","",COUNTIF(SubParent,${c_id}{r}))')
        ws.cell(row=r, column=L.T_EFFORT,
                value=f'=IF(${c_id}{r}="","",ROUND({effort},2))').number_format = "0.00"
        ws.cell(row=r, column=L.T_REMAINING,
                value=f'=IF(${c_id}{r}="","",ROUND(SUMIF(SubParent,${c_id}{r},SubRemaining),2))'
                ).number_format = "0.00"
        # Derived the same way effort is: a task's own status column would be a
        # second source of truth that can disagree with what its sub-tasks say.
        # Status *names* are configurable, so the label is looked up by position
        # rather than hard-coded, matching how Config documents the convention.
        done_label = f'INDEX(StatusNames,{L.STATUS_DONE})'
        todo_label = f'INDEX(StatusNames,{L.STATUS_TODO})'
        active_label = f'INDEX(StatusNames,{L.STATUS_ACTIVE})'
        ws.cell(row=r, column=L.T_STATUS, value=(
            f'=IF(${c_id}{r}="","",IF(${c_subs}{r}=0,"",'
            f'IF(COUNTIFS(SubParent,${c_id}{r},SubStatus,{done_label})=${c_subs}{r},{done_label},'
            f'IF(COUNTIFS(SubParent,${c_id}{r},SubStatus,{todo_label})=${c_subs}{r},{todo_label},'
            f'{active_label}))))'))
        # Start and end are resolved as grid positions and then rendered through
        # the shared label, so a task running into next year reads WW01 '27
        # rather than a week number that does not exist.
        start_pos = f'SUMPRODUCT(MIN(({span}>0)*CwPos+({span}<=0)*9999))'
        end_pos = f'SUMPRODUCT(MAX(({span}>0)*CwPos))'
        ws.cell(row=r, column=L.T_CALC_START, value=(
            f'=IF(${c_id}{r}="","",IF(SUM({span})=0,"",'
            f'IFERROR(INDEX(CwLabel,{start_pos}),"")))'))
        ws.cell(row=r, column=L.T_CALC_END, value=(
            f'=IF(${c_id}{r}="","",IF(SUM({span})=0,"",'
            f'IFERROR(INDEX(CwLabel,{end_pos}),"")))'))

        # A start week is valid only if it names one of the built columns *and*
        # that column is inside the horizon.
        pos = f'IFERROR(MATCH(${c_start}{r},CwWeeks,0),0)'
        ws.cell(row=r, column=L.T_CHECK, value=(
            f'=IF(${c_id}{r}="","",'
            f'IF(COUNTIF(SubParent,${c_id}{r})=0,"⚠ no sub-tasks",'
            # Finished work is checked first and reported as finished. It has no
            # remaining effort, so every test below would flag it — "not
            # scheduled" for a task that is simply done, which is the complaint
            # this phase exists to answer.
            f'IF(${c_status}{r}=INDEX(StatusNames,{L.STATUS_DONE}),"done",'
            f'IF({pos}=0,"⚠ start week not in the grid",'
            f'IF({pos}>CfgHorizon,"⚠ starts beyond the horizon",'
            f'IF(ROUND(${c_remaining}{r},4)<=0,"in progress",'
            f'IF(SUM({span})=0,"⚠ not scheduled",'
            f'IF(ROUND(SUM({span}),4)<ROUND({effort},4),"⚠ overruns horizon","ok"))))))))'))

        for c in L.TASK_DERIVED_COLS:
            S.mark_derived(ws.cell(row=r, column=c))
        for c in range(1, L.T_CHECK + 1):
            ws.cell(row=r, column=c).border = S.BORDER_ALL
        ws.cell(row=r, column=L.T_START_WW).number_format = '"WW"0'

    _list_validation(ws, "PrioNames", L.T_PRIORITY, L.TASK_FIRST_ROW, N.LAST_TASK_ROW)
    _list_validation(ws, "CplxNames", L.T_COMPLEXITY, L.TASK_FIRST_ROW, N.LAST_TASK_ROW)
    _list_validation(ws, "EqpNames", L.T_EQUIPMENT, L.TASK_FIRST_ROW, N.LAST_TASK_ROW)
    _list_validation(ws, "AsgNames", L.T_DEF_ASSIGNEE, L.TASK_FIRST_ROW, N.LAST_TASK_ROW)
    _check_formatting(ws, L.T_CHECK, L.TASK_FIRST_ROW, N.LAST_TASK_ROW)

    S.widths(ws, {"A": 10, "B": 24, "C": 12, "D": 9, "E": 12, "F": 15, "G": 17,
                  "H": 17, "I": 12, "J": 13, "K": 13, "L": 14, "M": 10, "N": 10,
                  "O": 22})
    ws.freeze_panes = "C2"
    r = N.LAST_TASK_ROW + 2
    ws.cell(row=r, column=1,
            value="Grey italic columns are computed — do not type in them. "
                  "Effort is the sum of the task's sub-tasks; task complexity is metadata.").font = S.FONT_NOTE


# --- Sub-Tasks -------------------------------------------------------------

def build_subtasks(ws) -> None:
    S.header_row(ws, 1, 1, [L.SUB_HEADERS[c] for c in sorted(L.SUB_HEADERS)])
    for c in L.SUB_DERIVED_COLS:
        ws.cell(row=1, column=c).fill = S.FILL_OUTPUT_HDR

    for i, sub in enumerate(demo.subtasks()):
        r = L.SUB_FIRST_ROW + i
        ws.cell(row=r, column=L.S_PARENT, value=sub["parent"])
        ws.cell(row=r, column=L.S_NAME, value=sub["name"])
        ws.cell(row=r, column=L.S_COMPLEXITY, value=sub["complexity"])
        if sub["assignee"]:
            ws.cell(row=r, column=L.S_ASSIGNEE, value=sub["assignee"])
        ws.cell(row=r, column=L.S_STATUS, value=sub["status"])
        ws.cell(row=r, column=L.S_PCT_DONE, value=sub["pct_done"])
        if sub.get("actual_start_week"):
            cell = ws.cell(row=r, column=L.S_ACT_START, value=sub["actual_start_week"])
            cell.number_format = '"WW"0'
        if sub.get("actual_end_week"):
            cell = ws.cell(row=r, column=L.S_ACT_END, value=sub["actual_end_week"])
            cell.number_format = '"WW"0'

    for r in range(L.SUB_FIRST_ROW, N.LAST_SUB_ROW + 1):
        c_parent = L.col(L.S_PARENT)
        c_cplx = L.col(L.S_COMPLEXITY)
        c_asg = L.col(L.S_ASSIGNEE)
        c_status = L.col(L.S_STATUS)
        c_pct = L.col(L.S_PCT_DONE)
        c_eff_asg = L.col(L.S_EFF_ASSIGNEE)
        c_effort = L.col(L.S_EFFORT)
        c_remaining = L.col(L.S_REMAINING)
        c_act_start = L.col(L.S_ACT_START)
        c_key = L.col(L.S_KEY)
        idx = f"COUNTIF(${c_parent}${L.SUB_FIRST_ROW}:${c_parent}{r},${c_parent}{r})"
        parent_row = f"MATCH(${c_parent}{r},TaskIDs,0)"
        ws.cell(row=r, column=L.S_ID,
                value=f'=IF(${c_parent}{r}="","",${c_parent}{r}&"."&TEXT({idx},"00"))')
        # `&""` matters: INDEX on an empty default-assignee cell returns 0, which
        # would display as a literal 0 rather than reading as "nobody set yet".
        ws.cell(row=r, column=L.S_EFF_ASSIGNEE, value=(
            f'=IF(${c_parent}{r}="","",IF(${c_asg}{r}<>"",${c_asg}{r},'
            f'IFERROR(INDEX(TaskDefAsg,{parent_row})&"","")))'))
        ws.cell(row=r, column=L.S_EFFORT, value=(
            f'=IF(OR(${c_parent}{r}="",${c_cplx}{r}="",${c_eff_asg}{r}=""),"",'
            f'IFERROR(INDEX(CplxDays,MATCH(${c_cplx}{r},CplxNames,0))'
            f'/INDEX(AsgProf,MATCH(${c_eff_asg}{r},AsgNames,0)),""))')).number_format = "0.00"
        # Done claims no capacity regardless of what % done says; otherwise the
        # balance is what remains of effort after the recorded progress.
        ws.cell(row=r, column=L.S_REMAINING, value=(
            f'=IF(${c_effort}{r}="","",IF(IFERROR(MATCH(${c_status}{r},StatusNames,0),0)={L.STATUS_DONE},0,'
            f'${c_effort}{r}*(1-N(${c_pct}{r})/100)))')).number_format = "0.00"
        ws.cell(row=r, column=L.S_CONSUMED, value=(
            f'=IF(${c_effort}{r}="","",'
            f'ROUND(${c_effort}{r}-${c_remaining}{r},4))')).number_format = "0.00"
        ws.cell(row=r, column=L.S_KEY, value=(
            f'=IF(${c_effort}{r}="","",IFERROR('
            f'INDEX(PrioRank,MATCH(INDEX(TaskPrio,{parent_row}),PrioNames,0))*1000000000'
            f'+INDEX(TaskStartWW,{parent_row})*1000000'
            f'+{parent_row}*1000+{idx},""))')).number_format = "0"
        ws.cell(row=r, column=L.S_RANK,
                value=f'=IF(${c_key}{r}="","",RANK(${c_key}{r},SubKey,1))')
        ws.cell(row=r, column=L.S_CHECK, value=(
            f'=IF(${c_parent}{r}="","",'
            f'IF(ISNA({parent_row}),"⚠ unknown parent",'
            f'IF(${c_cplx}{r}="","⚠ no complexity",'
            f'IF(${c_eff_asg}{r}="","⚠ no assignee",'
            f'IF(${c_key}{r}="","⚠ check proficiency",'
            f'IF(AND(IFERROR(MATCH(${c_status}{r},StatusNames,0),0)>{L.STATUS_TODO},'
            f'${c_act_start}{r}=""),"⚠ no actual start",'
            f'IF(IFERROR(MATCH(${c_status}{r},StatusNames,0),0)={L.STATUS_DONE},'
            f'"done","ok")))))))'))

        for c in L.SUB_DERIVED_COLS:
            S.mark_derived(ws.cell(row=r, column=c))
        for c in range(1, L.S_CHECK + 1):
            ws.cell(row=r, column=c).border = S.BORDER_ALL

    _list_validation(ws, "TaskIDs", L.S_PARENT, L.SUB_FIRST_ROW, N.LAST_SUB_ROW)
    _list_validation(ws, "CplxNames", L.S_COMPLEXITY, L.SUB_FIRST_ROW, N.LAST_SUB_ROW)
    _list_validation(ws, "AsgNames", L.S_ASSIGNEE, L.SUB_FIRST_ROW, N.LAST_SUB_ROW)
    _list_validation(ws, "StatusNames", L.S_STATUS, L.SUB_FIRST_ROW, N.LAST_SUB_ROW)
    _check_formatting(ws, L.S_CHECK, L.SUB_FIRST_ROW, N.LAST_SUB_ROW)
    _parent_looks_merged(ws)

    S.widths(ws, {"A": 12, "B": 12, "C": 26, "D": 12, "E": 22, "F": 14,
                  "G": 10, "H": 18, "I": 12, "J": 14, "K": 16, "L": 8, "M": 22})
    ws.freeze_panes = "C2"
    r = N.LAST_SUB_ROW + 2
    ws.cell(row=r, column=1,
            value="Leave Assignee blank to inherit the parent task's default assignee. "
                  "Repeated parent IDs are hidden for readability — the value is still "
                  "in every cell, so sorting and filtering keep working.").font = S.FONT_NOTE


def _parent_looks_merged(ws) -> None:
    """Render repeated parent IDs as one block without actually merging.

    A real merge would store the value only in the top-left cell and would break
    sort and AutoFilter on a table the user edits. Instead every row keeps its
    real parent ID and the repeats are drawn in white.
    """
    first, last = L.SUB_FIRST_ROW, N.LAST_SUB_ROW
    parent_col = L.col(L.S_PARENT)

    repeats = f"{parent_col}{first}:{parent_col}{last}"
    ws.conditional_formatting.add(repeats, Rule(
        type="expression",
        dxf=DifferentialStyle(font=Font(color=S.cf_color("FFFFFF"))),
        formula=[f'AND(${parent_col}{first}<>"",${parent_col}{first}=${parent_col}{first - 1})']))

    whole_row = f"A{first}:{L.col(L.S_CHECK)}{last}"
    ws.conditional_formatting.add(whole_row, Rule(
        type="expression",
        dxf=DifferentialStyle(border=Border(top=Side(style="medium", color=S.cf_color(S.INK)))),
        formula=[f'AND(${parent_col}{first}<>"",${parent_col}{first}<>${parent_col}{first - 1})']))
