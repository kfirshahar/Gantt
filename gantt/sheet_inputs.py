"""The tabs the user types into: Config, Assignees, Capacity, Equipment, Holidays."""

from datetime import datetime

from openpyxl.styles import Alignment

from . import demo, layout as L, names as N, styles as S


def _week_headers(ws, row: int, first_col: int, count: int, output: bool = False) -> None:
    """Week-number header cells driven by Config, so shifting the project
    start week relabels every grid in the workbook at once."""
    for i in range(count):
        cell = ws.cell(row=row, column=first_col + i, value=f"=CfgStartWeek+{i}")
        cell.number_format = '"WW"0'
        cell.fill = S.FILL_OUTPUT_HDR if output else S.FILL_INPUT_HDR
        cell.font = S.FONT_HDR
        cell.alignment = S.CENTER
        cell.border = S.BORDER_ALL
        ws.column_dimensions[L.col(first_col + i)].width = 8


def build_config(ws) -> None:
    S.title(ws, "Configuration",
            "Duplicate this workbook and edit the values below to fit another project.")

    rows = [
        (L.CFG_YEAR_ROW, "Project year", demo.YEAR, "0"),
        (L.CFG_START_WEEK_ROW, "Start week (WW)", demo.START_WEEK, "0"),
        (L.CFG_HORIZON_ROW, "Horizon (weeks)", demo.HORIZON, "0"),
    ]
    for row, label, value, fmt in rows:
        ws.cell(row=row, column=1, value=label).font = S.FONT_SECTION
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.border = S.BORDER_ALL

    note = ws.cell(row=L.CFG_HORIZON_ROW, column=3,
                   value=f"Grids are pre-built for {L.HORIZON_WEEKS} weeks.")
    note.font = S.FONT_NOTE

    S.header_row(ws, L.CFG_COMPLEXITY_HDR, 1, ["Complexity", "Base days"])
    for i, (name, days) in enumerate(demo.COMPLEXITY):
        r = L.CFG_COMPLEXITY_FIRST + i
        ws.cell(row=r, column=1, value=name).border = S.BORDER_ALL
        c = ws.cell(row=r, column=2, value=days)
        c.number_format = "0.0"
        c.border = S.BORDER_ALL

    S.header_row(ws, L.CFG_PRIORITY_HDR, 1, ["Priority", "Order"])
    for i, (name, order) in enumerate(demo.PRIORITIES):
        r = L.CFG_PRIORITY_FIRST + i
        ws.cell(row=r, column=1, value=name).border = S.BORDER_ALL
        c = ws.cell(row=r, column=2, value=order)
        c.border = S.BORDER_ALL

    ws.cell(row=L.CFG_PRIORITY_FIRST + L.CFG_PRIORITY_COUNT + 1, column=1,
            value="Lower 'Order' claims capacity first.").font = S.FONT_NOTE
    S.widths(ws, {"A": 22, "B": 12, "C": 44})


def build_assignees(ws) -> None:
    S.header_row(ws, 1, 1, ["Name", "Proficiency", "Notes"])
    for i, (name, prof) in enumerate(demo.ASSIGNEES):
        r = L.GRID_FIRST_DATA_ROW + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=prof).number_format = "0.00"
    for r in range(L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW + 1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = S.BORDER_ALL
    S.widths(ws, {"A": 18, "B": 12, "C": 40})
    ws.freeze_panes = "A2"
    ws.cell(row=N.LAST_ASG_ROW + 2, column=1,
            value="Proficiency scales effort: at 1.2 a job takes 1/1.2 of its base days, "
                  "at 0.8 it takes 1/0.8. Leave at 1.00 for an average performer.").font = S.FONT_NOTE


def _linked_name_column(ws, source_sheet: str, last_row: int) -> None:
    """Mirror the Assignees/Equipment name list so grids stay in step."""
    for r in range(L.GRID_FIRST_DATA_ROW, last_row + 1):
        cell = ws.cell(
            row=r, column=1,
            value=f'=IF({L.sheet_ref(source_sheet)}!A{r}="","",{L.sheet_ref(source_sheet)}!A{r})')
        S.mark_derived(cell)
        cell.border = S.BORDER_ALL


def build_capacity(ws) -> None:
    S.header_row(ws, 1, 1, ["Assignee"])
    _week_headers(ws, 1, L.GRID_FIRST_WEEK_COL, L.HORIZON_WEEKS)
    _linked_name_column(ws, L.ASSIGNEES, N.LAST_ASG_ROW)

    by_name = {name: demo.CAPACITY.get(name, []) for name, _ in demo.ASSIGNEES}
    for i, (name, _) in enumerate(demo.ASSIGNEES):
        r = L.GRID_FIRST_DATA_ROW + i
        for w in range(L.HORIZON_WEEKS):
            values = by_name[name]
            if w < len(values):
                ws.cell(row=r, column=L.GRID_FIRST_WEEK_COL + w, value=values[w])

    total_col = N.LAST_WEEK_COL + 1
    S.header_row(ws, 1, total_col, ["Total"], output=True)
    for r in range(L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW + 1):
        for c in range(L.GRID_FIRST_WEEK_COL, N.LAST_WEEK_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = "0.0"
            cell.alignment = S.CENTER
            cell.border = S.BORDER_ALL
        total = ws.cell(row=r, column=total_col, value=(
            f'=IF($A{r}="","",SUM({L.col(L.GRID_FIRST_WEEK_COL)}{r}:'
            f'{L.col(N.LAST_WEEK_COL)}{r}))'))
        total.number_format = "0.0"
        total.alignment = S.CENTER
        total.border = S.BORDER_ALL
        S.mark_derived(total)

    S.widths(ws, {"A": 18, L.col(total_col): 9})
    ws.freeze_panes = "B2"
    r = N.LAST_ASG_ROW + 2
    ws.cell(row=r, column=1,
            value="Available work days per assignee per week "
                  "(holidays, vacation and part-time bandwidth already netted out).").font = S.FONT_NOTE


def build_equipment(ws) -> None:
    S.header_row(ws, 1, 1, ["Equipment type"])
    _week_headers(ws, 1, L.GRID_FIRST_WEEK_COL, L.HORIZON_WEEKS)

    for i, (name, units) in enumerate(demo.EQUIPMENT.items()):
        r = L.GRID_FIRST_DATA_ROW + i
        ws.cell(row=r, column=1, value=name)
        for w in range(min(L.HORIZON_WEEKS, len(units))):
            ws.cell(row=r, column=L.GRID_FIRST_WEEK_COL + w, value=units[w])

    for r in range(L.GRID_FIRST_DATA_ROW, N.LAST_EQP_ROW + 1):
        ws.cell(row=r, column=1).border = S.BORDER_ALL
        for c in range(L.GRID_FIRST_WEEK_COL, N.LAST_WEEK_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = "0"
            cell.alignment = S.CENTER
            cell.border = S.BORDER_ALL

    S.widths(ws, {"A": 18})
    ws.freeze_panes = "B2"
    r = N.LAST_EQP_ROW + 2
    ws.cell(row=r, column=1,
            value="Units available in the shared pool each week. "
                  "Shortages are flagged on Gantt-High; they do not reschedule work.").font = S.FONT_NOTE


def build_holidays(ws) -> None:
    S.header_row(ws, 1, 1, ["Date", "Name"])
    for i, (iso, name) in enumerate(demo.HOLIDAYS):
        r = L.GRID_FIRST_DATA_ROW + i
        cell = ws.cell(row=r, column=1, value=datetime.strptime(iso, "%Y-%m-%d"))
        cell.number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=2, value=name)
    for r in range(L.GRID_FIRST_DATA_ROW, N.LAST_HOL_ROW + 1):
        for c in (1, 2):
            ws.cell(row=r, column=c).border = S.BORDER_ALL
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    S.widths(ws, {"A": 14, "B": 30})
    r = N.LAST_HOL_ROW + 2
    ws.cell(row=r, column=1,
            value="Company-wide non-working days. The demo dates are placeholders — "
                  "replace them with your own calendar.").font = S.FONT_NOTE
