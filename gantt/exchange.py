"""Moving the input data in and out of a workbook as JSON.

Phase 1 of the roadmap: export only.

One invariant shapes everything here — **nothing computed is ever read**. Every
value comes from a cell the user types into, which holds a literal. That matters
because openpyxl does not evaluate formulas: a workbook generated a moment ago
has no cached results at all, so an exporter that reached for a derived column
would return `None` for a freshly built file and a stale number for an edited
one. Derived facts the JSON needs, such as a sub-task's ID, are recomputed here
using the same rule the workbook uses.

The second consequence is that export needs no recalculation engine, so it runs
anywhere.

Week-indexed grids are exported by **offset from `config.start_week`**, not by
calendar week. Calendar labels live in computed cells, and a positional list
plus the start week says the same thing without depending on them.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import layout as L, names as N

SCHEMA_VERSION = 1


def _clean(value: Any) -> Any:
    """A cell's value, with blanks normalised to None."""
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _row_is_empty(ws, row: int, columns) -> bool:
    return all(_clean(ws.cell(row=row, column=c).value) is None for c in columns)


def _to_excel_date(value: Any) -> Any:
    """Turn an ISO date string back into a datetime.

    Export renders dates as ISO strings, and writing one back verbatim leaves a
    cell holding text. Excel compares text against a date serial without ever
    matching, so every holiday would be silently ignored and every week would
    look like a full five days. Nothing errors; the plan is just quietly wrong.
    """
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d")
        except ValueError:
            return value
    return value


def _as_iso(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip() or None
    return None


def _config(wb) -> dict:
    ws = wb[L.CONFIG]
    complexity = []
    for i in range(L.CFG_COMPLEXITY_COUNT):
        row = L.CFG_COMPLEXITY_FIRST + i
        name = _clean(ws.cell(row=row, column=1).value)
        if name is not None:
            complexity.append({"name": name,
                               "base_days": _clean(ws.cell(row=row, column=2).value)})

    priorities = []
    for i in range(L.CFG_PRIORITY_COUNT):
        row = L.CFG_PRIORITY_FIRST + i
        name = _clean(ws.cell(row=row, column=1).value)
        if name is not None:
            priorities.append({"name": name,
                               "order": _clean(ws.cell(row=row, column=2).value)})

    return {
        "year": _clean(ws.cell(row=L.CFG_YEAR_ROW, column=2).value),
        "start_week": _clean(ws.cell(row=L.CFG_START_WEEK_ROW, column=2).value),
        "horizon_weeks": _clean(ws.cell(row=L.CFG_HORIZON_ROW, column=2).value),
        "complexity": complexity,
        "priorities": priorities,
    }


def _assignees(wb) -> list[dict]:
    ws = wb[L.ASSIGNEES]
    out = []
    for row in range(L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW + 1):
        name = _clean(ws.cell(row=row, column=1).value)
        if name is None:
            continue
        out.append({"name": name,
                    "proficiency": _clean(ws.cell(row=row, column=2).value),
                    "notes": _clean(ws.cell(row=row, column=3).value)})
    return out


def _week_grid(wb, sheet: str, labels: list[str], key: str, values_key: str) -> list[dict]:
    """A per-week grid, exported by offset from the project start week.

    `labels` supplies the row identities: the Capacity sheet's own name column is
    a formula mirroring Assignees, so it cannot be read, and the row order is
    what ties the two together.
    """
    ws = wb[sheet]
    out = []
    for i, label in enumerate(labels):
        row = L.GRID_FIRST_DATA_ROW + i
        series = [_clean(ws.cell(row=row, column=L.GRID_FIRST_WEEK_COL + w).value)
                  for w in range(L.WEEK_COLS)]
        out.append({key: label, values_key: series})
    return out


def _equipment_names(wb) -> list[str]:
    ws = wb[L.EQUIPMENT]
    names = []
    for row in range(L.GRID_FIRST_DATA_ROW, N.LAST_EQP_ROW + 1):
        name = _clean(ws.cell(row=row, column=1).value)
        if name is not None:
            names.append(name)
    return names


def _holidays(wb) -> list[dict]:
    ws = wb[L.HOLIDAYS]
    out = []
    for row in range(L.GRID_FIRST_DATA_ROW, N.LAST_HOL_ROW + 1):
        when = _as_iso(ws.cell(row=row, column=1).value)
        if when is None:
            continue
        out.append({"date": when, "name": _clean(ws.cell(row=row, column=2).value)})
    return out


def _tasks(wb) -> list[dict]:
    ws = wb[L.TASKS]
    out = []
    for row in range(L.TASK_FIRST_ROW, N.LAST_TASK_ROW + 1):
        if _row_is_empty(ws, row, L.TASK_INPUT_COLS):
            continue
        out.append({
            "id": _clean(ws.cell(row=row, column=L.T_ID).value),
            "name": _clean(ws.cell(row=row, column=L.T_NAME).value),
            "category": _clean(ws.cell(row=row, column=L.T_CATEGORY).value),
            "priority": _clean(ws.cell(row=row, column=L.T_PRIORITY).value),
            "complexity": _clean(ws.cell(row=row, column=L.T_COMPLEXITY).value),
            "equipment": _clean(ws.cell(row=row, column=L.T_EQUIPMENT).value),
            "default_assignee": _clean(ws.cell(row=row, column=L.T_DEF_ASSIGNEE).value),
            "earliest_start_week": _clean(ws.cell(row=row, column=L.T_START_WW).value),
        })
    return out


def _sub_tasks(wb) -> list[dict]:
    """Sub-tasks, with the ID recomputed rather than read.

    The workbook builds the ID with a formula, so the cell holds no value until
    something recalculates. The rule is simply the parent plus a running count
    of that parent from the top, which is cheap to reproduce — and reproducing it
    keeps export working on a file that has never been opened in Excel.
    """
    ws = wb[L.SUBTASKS]
    seen: dict[str, int] = {}
    out = []
    for row in range(L.SUB_FIRST_ROW, N.LAST_SUB_ROW + 1):
        if _row_is_empty(ws, row, L.SUB_INPUT_COLS):
            continue
        parent = _clean(ws.cell(row=row, column=L.S_PARENT).value)
        index = None
        if parent is not None:
            seen[parent] = seen.get(parent, 0) + 1
            index = seen[parent]
        out.append({
            "id": f"{parent}.{index:02d}" if parent is not None else None,
            "parent": parent,
            "name": _clean(ws.cell(row=row, column=L.S_NAME).value),
            "complexity": _clean(ws.cell(row=row, column=L.S_COMPLEXITY).value),
            "assignee": _clean(ws.cell(row=row, column=L.S_ASSIGNEE).value),
        })
    return out


def export(path: str | Path) -> dict:
    """Every input tab of a workbook, as a plain dictionary."""
    wb = load_workbook(Path(path), data_only=False)
    assignees = _assignees(wb)

    return {
        "schema_version": SCHEMA_VERSION,
        "source": Path(path).name,
        "config": _config(wb),
        "assignees": assignees,
        "capacity": _week_grid(wb, L.CAPACITY, [a["name"] for a in assignees],
                               "assignee", "days_by_week_offset"),
        "equipment": _week_grid(wb, L.EQUIPMENT, _equipment_names(wb),
                                "type", "units_by_week_offset"),
        "holidays": _holidays(wb),
        "tasks": _tasks(wb),
        "sub_tasks": _sub_tasks(wb),
    }


def to_json(data: dict) -> str:
    return json.dumps(data, indent=2, ensure_ascii=False, sort_keys=False)


# --- Import ----------------------------------------------------------------
#
# Two modes, because first ingest and a weekly update want opposite things.
#
# `replace` treats the JSON as the whole truth and is what a first ingest or a
# rebuild onto a new template version needs.
#
# `merge` respects field ownership: the JSON owns facts observed in the outside
# world, the workbook owns planning decisions. Without that rule every weekly
# import would silently revert whatever the planner had changed since the last
# one. A row that does not exist yet is written in full either way — ownership
# only governs what an *update* may overwrite.

TASK_FIELD_COL = {
    "id": L.T_ID, "name": L.T_NAME, "category": L.T_CATEGORY,
    "priority": L.T_PRIORITY, "complexity": L.T_COMPLEXITY,
    "equipment": L.T_EQUIPMENT, "default_assignee": L.T_DEF_ASSIGNEE,
    "earliest_start_week": L.T_START_WW,
}
SUB_FIELD_COL = {
    "parent": L.S_PARENT, "name": L.S_NAME,
    "complexity": L.S_COMPLEXITY, "assignee": L.S_ASSIGNEE,
}

# Descriptive facts come from the source system; the rest are the planner's.
TASK_JSON_OWNED = {"name", "category"}
SUB_JSON_OWNED = {"name"}


class Change:
    """One cell the import intends to write. Collected before anything is."""

    __slots__ = ("sheet", "row", "column", "field", "old", "new", "reason")

    def __init__(self, sheet, row, column, field, old, new, reason):
        self.sheet, self.row, self.column = sheet, row, column
        self.field, self.old, self.new, self.reason = field, old, new, reason

    def __repr__(self) -> str:
        return (f"{self.sheet}!{L.col(self.column)}{self.row} {self.field}: "
                f"{self.old!r} -> {self.new!r} ({self.reason})")


def _set(changes, ws, row, column, field, value, reason):
    old = _clean(ws.cell(row=row, column=column).value)
    new = _clean(value)
    if old != new:
        changes.append(Change(ws.title, row, column, field, old, new, reason))


def _plan_rows(changes, ws, rows, field_col, first_row, last_row,
               key_of, existing_keys, json_owned, mode) -> set:
    """Update matched rows and append unmatched ones.

    Returns the rows this import actually touched. `existing_keys` describes the
    workbook as it stands, so it is emphatically not that set: a row present in
    the file but absent from the JSON must be cleared under replace, and using
    the wrong set here leaves demo data mixed into freshly imported data.
    """
    free = [r for r in range(first_row, last_row + 1)
            if r not in existing_keys.values()
            and _row_is_empty(ws, r, list(field_col.values()))]

    used = set()
    for record in rows:
        key = key_of(record)
        row = existing_keys.get(key)
        inserting = row is None
        if inserting:
            if not free:
                raise ValueError(
                    f"{ws.title} is full: no free row for {key!r}. The template "
                    f"pre-builds {last_row - first_row + 1} rows; raise the limit "
                    f"in gantt/layout.py and rebuild.")
            row = free.pop(0)
            existing_keys[key] = row
        used.add(row)

        for field, column in field_col.items():
            if field not in record:
                continue
            # Ownership only constrains updates; a new row is written in full.
            if mode == "merge" and not inserting and field not in json_owned:
                continue
            _set(changes, ws, row, column, field, record[field],
                 "insert" if inserting else "update")
    return used


def _plan_clear(changes, ws, field_col, rows_used, first_row, last_row):
    for row in range(first_row, last_row + 1):
        if row in rows_used:
            continue
        for field, column in field_col.items():
            _set(changes, ws, row, column, field, None, "clear")


def _plan_tasks(changes, wb, data, mode):
    ws = wb[L.TASKS]
    existing = {}
    for row in range(L.TASK_FIRST_ROW, N.LAST_TASK_ROW + 1):
        tid = _clean(ws.cell(row=row, column=L.T_ID).value)
        if tid is not None:
            existing[tid] = row

    used = _plan_rows(changes, ws, data.get("tasks", []), TASK_FIELD_COL,
                      L.TASK_FIRST_ROW, N.LAST_TASK_ROW,
                      lambda r: r["id"], existing, TASK_JSON_OWNED, mode)
    if mode == "replace":
        _plan_clear(changes, ws, TASK_FIELD_COL, used,
                    L.TASK_FIRST_ROW, N.LAST_TASK_ROW)


def _plan_sub_tasks(changes, wb, data, mode):
    """Matched on (parent, name), not on the ID.

    Sub-task IDs are positional — T-01.01 is simply the first T-01 row from the
    top — so inserting a row silently renames every one below it. Matching on
    them would attach an update to the wrong work. The pair is stable under
    insertion and reordering, and names are unique within a parent in practice.
    """
    ws = wb[L.SUBTASKS]
    existing = {}
    for row in range(L.SUB_FIRST_ROW, N.LAST_SUB_ROW + 1):
        parent = _clean(ws.cell(row=row, column=L.S_PARENT).value)
        name = _clean(ws.cell(row=row, column=L.S_NAME).value)
        if parent is not None:
            existing[(parent, name)] = row

    used = _plan_rows(changes, ws, data.get("sub_tasks", []), SUB_FIELD_COL,
                      L.SUB_FIRST_ROW, N.LAST_SUB_ROW,
                      lambda r: (r["parent"], r.get("name")), existing,
                      SUB_JSON_OWNED, mode)
    if mode == "replace":
        _plan_clear(changes, ws, SUB_FIELD_COL, used,
                    L.SUB_FIRST_ROW, N.LAST_SUB_ROW)


def _plan_reference(changes, wb, data, mode):
    """Assignees, Equipment and Holidays.

    On merge these are the planner's to own, so existing entries are never
    rewritten — but a missing one is appended anyway. A new task naming an
    assignee the workbook has never heard of would otherwise fail its dropdown
    and schedule nothing, which is a worse outcome than adding the row.
    """
    ws = wb[L.ASSIGNEES]
    fields = {"name": 1, "proficiency": 2, "notes": 3}
    existing = {}
    for row in range(L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW + 1):
        name = _clean(ws.cell(row=row, column=1).value)
        if name is not None:
            existing[name] = row
    used = _plan_rows(changes, ws, data.get("assignees", []), fields,
                      L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW,
                      lambda r: r["name"], existing, set(), mode)
    if mode == "replace":
        _plan_clear(changes, ws, fields, used,
                    L.GRID_FIRST_DATA_ROW, N.LAST_ASG_ROW)

    ws = wb[L.EQUIPMENT]
    existing = {}
    for row in range(L.GRID_FIRST_DATA_ROW, N.LAST_EQP_ROW + 1):
        name = _clean(ws.cell(row=row, column=1).value)
        if name is not None:
            existing[name] = row
    used = _plan_rows(changes, ws, data.get("equipment", []), {"type": 1},
                      L.GRID_FIRST_DATA_ROW, N.LAST_EQP_ROW,
                      lambda r: r["type"], existing, set(), mode)
    if mode == "replace":
        _plan_clear(changes, ws, {"type": 1}, used,
                    L.GRID_FIRST_DATA_ROW, N.LAST_EQP_ROW)

    if mode == "replace":
        ws = wb[L.HOLIDAYS]
        rows = data.get("holidays", [])
        for i in range(L.MAX_HOLIDAYS):
            row = L.GRID_FIRST_DATA_ROW + i
            record = rows[i] if i < len(rows) else {}
            _set(changes, ws, row, 1, "date",
                 _to_excel_date(record.get("date")), "replace")
            _set(changes, ws, row, 2, "name", record.get("name"), "replace")


def _plan_week_grids(changes, wb, data, mode):
    """Capacity and equipment availability. Planning data, so merge never
    touches them; the planner decides who is available when."""
    if mode != "replace":
        return

    # Every row in range is written, not just the supplied ones: a row left
    # untouched keeps whatever numbers were there before, and a stale capacity
    # figure under a name that has been cleared is invisible but still counted.
    order = [a["name"] for a in data.get("assignees", [])]
    capacity = {c["assignee"]: c.get("days_by_week_offset", [])
                for c in data.get("capacity", [])}
    for i in range(L.MAX_ASSIGNEES):
        series = capacity.get(order[i], []) if i < len(order) else []
        row = L.GRID_FIRST_DATA_ROW + i
        for w in range(L.WEEK_COLS):
            value = series[w] if w < len(series) else None
            _set(changes, wb[L.CAPACITY], row, L.GRID_FIRST_WEEK_COL + w,
                 f"week+{w}", value, "replace")

    equipment = data.get("equipment", [])
    for i in range(L.MAX_EQUIPMENT):
        series = equipment[i].get("units_by_week_offset", []) if i < len(equipment) else []
        row = L.GRID_FIRST_DATA_ROW + i
        for w in range(L.WEEK_COLS):
            value = series[w] if w < len(series) else None
            _set(changes, wb[L.EQUIPMENT], row, L.GRID_FIRST_WEEK_COL + w,
                 f"week+{w}", value, "replace")


def _plan_config(changes, wb, data, mode):
    if mode != "replace":
        return
    ws, config = wb[L.CONFIG], data.get("config", {})
    for row, field in ((L.CFG_YEAR_ROW, "year"),
                       (L.CFG_START_WEEK_ROW, "start_week"),
                       (L.CFG_HORIZON_ROW, "horizon_weeks")):
        if config.get(field) is not None:
            _set(changes, ws, row, 2, field, config[field], "replace")

    for i in range(L.CFG_COMPLEXITY_COUNT):
        row = L.CFG_COMPLEXITY_FIRST + i
        entry = config.get("complexity", [])
        record = entry[i] if i < len(entry) else {}
        _set(changes, ws, row, 1, "complexity", record.get("name"), "replace")
        _set(changes, ws, row, 2, "base_days", record.get("base_days"), "replace")

    for i in range(L.CFG_PRIORITY_COUNT):
        row = L.CFG_PRIORITY_FIRST + i
        entry = config.get("priorities", [])
        record = entry[i] if i < len(entry) else {}
        _set(changes, ws, row, 1, "priority", record.get("name"), "replace")
        _set(changes, ws, row, 2, "order", record.get("order"), "replace")


def plan(wb, data: dict, mode: str = "merge") -> list[Change]:
    """Every cell the import would write, without writing any of them."""
    if mode not in ("merge", "replace"):
        raise ValueError(f"unknown mode {mode!r}; expected 'merge' or 'replace'")

    version = data.get("schema_version")
    if version is not None and version > SCHEMA_VERSION:
        raise ValueError(
            f"this data is schema v{version} but the tool understands v"
            f"{SCHEMA_VERSION}. Upgrade the tool, or export again from the "
            f"version that wrote it.")

    changes: list[Change] = []
    _plan_config(changes, wb, data, mode)
    _plan_reference(changes, wb, data, mode)
    _plan_week_grids(changes, wb, data, mode)
    _plan_tasks(changes, wb, data, mode)
    _plan_sub_tasks(changes, wb, data, mode)
    return changes


def apply(wb, changes: list[Change]) -> None:
    for change in changes:
        wb[change.sheet].cell(row=change.row, column=change.column).value = change.new


def summarise(changes: list[Change]) -> str:
    if not changes:
        return "no changes"
    counts: dict[tuple[str, str], int] = {}
    for change in changes:
        counts[(change.sheet, change.reason)] = counts.get((change.sheet, change.reason), 0) + 1
    lines = [f"{len(changes)} cell(s) would change:"]
    for (sheet, reason), count in sorted(counts.items()):
        lines.append(f"  {sheet:12} {reason:8} {count}")
    return "\n".join(lines)


def import_data(workbook: str | Path, data: dict, mode: str = "merge",
                dry_run: bool = False, output: str | Path | None = None):
    """Apply JSON to a workbook. Returns the changes, applied unless dry_run."""
    source = Path(workbook)
    wb = load_workbook(source, data_only=False)
    changes = plan(wb, data, mode)
    if not dry_run:
        apply(wb, changes)
        wb.save(Path(output) if output else source)
    return changes


def rebuild(data: dict, output: str | Path):
    """A fresh workbook from JSON, using the current generator.

    This is the version-upgrade path: export from the old template, rebuild with
    the new one. Fields the old export never knew about simply take whatever
    default the generator gives them, so compatibility needs no special casing.
    """
    from .build import build

    target = Path(output)
    build(target)
    return import_data(target, data, mode="replace")


def _main(argv: list[str] | None = None) -> int:
    import argparse

    parser = argparse.ArgumentParser(
        prog="python -m gantt.exchange",
        description="Move a Gantt workbook's input data in and out as JSON.")
    sub = parser.add_subparsers(dest="command", required=True)

    exporter = sub.add_parser("export", help="write a workbook's input tabs as JSON")
    exporter.add_argument("workbook", type=Path)
    exporter.add_argument("-o", "--output", type=Path,
                          help="file to write; stdout if omitted")

    importer = sub.add_parser(
        "import", help="apply JSON to an existing workbook")
    importer.add_argument("workbook", type=Path)
    importer.add_argument("--json", dest="data", type=Path, required=True)
    importer.add_argument(
        "--mode", choices=("merge", "replace"), default="merge",
        help="merge respects field ownership (default); replace treats the "
             "JSON as the whole truth")
    importer.add_argument("--dry-run", action="store_true",
                          help="report what would change without writing")
    importer.add_argument("-o", "--output", type=Path,
                          help="write here instead of in place")
    importer.add_argument("-v", "--verbose", action="store_true")

    rebuilder = sub.add_parser(
        "rebuild", help="generate a fresh workbook from JSON")
    rebuilder.add_argument("data", type=Path)
    rebuilder.add_argument("-o", "--output", type=Path, required=True)

    args = parser.parse_args(argv)

    if args.command == "export":
        text = to_json(export(args.workbook))
        if args.output:
            args.output.write_text(text + "\n", encoding="utf-8")
            print(f"wrote {args.output}")
        else:
            print(text)

    elif args.command == "import":
        data = json.loads(args.data.read_text(encoding="utf-8"))
        changes = import_data(args.workbook, data, mode=args.mode,
                              dry_run=args.dry_run, output=args.output)
        print(summarise(changes))
        if args.verbose:
            for change in changes:
                print(f"  {change}")
        if args.dry_run:
            print("\ndry run: nothing was written")
        else:
            print(f"wrote {args.output or args.workbook}")

    elif args.command == "rebuild":
        data = json.loads(args.data.read_text(encoding="utf-8"))
        changes = rebuild(data, args.output)
        print(f"{summarise(changes)}\nwrote {args.output}")
    return 0

if __name__ == "__main__":
    raise SystemExit(_main())
