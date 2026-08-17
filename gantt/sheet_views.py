"""The read-only Gantt views."""

from openpyxl.formatting.rule import Rule
from openpyxl.styles import Border, Font, Side
from openpyxl.styles.differential import DifferentialStyle

from . import demo, layout as L, names as N, styles as S

# --- Gantt-High ------------------------------------------------------------

GH_WEEK_COL = 3          # weeks start at column C
GH_LOAD_HDR = 4
GH_LOAD_FIRST = 5


def build_gantt_high(ws) -> None:
    S.title(ws, "Gantt — high level",
            "Read-only. Weeks run across; every figure is in work days.")

    last_week_col = GH_WEEK_COL + L.WEEK_COLS - 1
    S.write_flag_row(ws, GH_WEEK_COL, L.WEEK_COLS, L.GH_FLAG_ROW,
                     lambda i: f"INDEX(CwActive,{i + 1})")
    # Which weeks are history. Conditional formatting has to read this from its
    # own sheet, and it is what tells actual apart from planned in every block.
    S.write_flag_row(
        ws, GH_WEEK_COL, L.WEEK_COLS, L.GH_HIST_ROW,
        lambda i: f"IF({i + 1}<IFERROR(MATCH(CfgCurrentWeek,CwWeeks,0),1),1,0)")
    load_last = _assignee_load(ws, last_week_col)
    timeline_last = _task_timeline(ws, load_last + 2, last_week_col)
    equip_last = _equipment_block(ws, timeline_last + 2, last_week_col)
    checks_last = _checks(ws, equip_last + 2)
    _plan_health(ws, checks_last + 2, last_week_col + 1)

    S.widths(ws, {"A": 38, "B": 9})
    for i in range(L.WEEK_COLS):
        ws.column_dimensions[L.col(GH_WEEK_COL + i)].width = 8
    ws.column_dimensions[L.col(last_week_col + 1)].width = 11
    ws.freeze_panes = L.col(GH_WEEK_COL) + str(GH_LOAD_FIRST)


def _week_header(ws, row: int, last_col: int, first_labels: list[str]) -> None:
    S.header_row(ws, row, 1, first_labels, output=True)
    for i in range(L.WEEK_COLS):
        cell = ws.cell(row=row, column=GH_WEEK_COL + i,
                       value=f"=INDEX(CwLabel,{i + 1})")
        cell.fill = S.FILL_OUTPUT_HDR
        cell.font = S.FONT_HDR
        cell.alignment = S.CENTER
        cell.border = S.BORDER_ALL


def _assignee_load(ws, last_col: int) -> int:
    """Two rows per assignee: days used, days available. Red where used > available."""
    S.section(ws, GH_LOAD_HDR - 1, "Assignee load — days used vs days available",
              span=last_col)
    _week_header(ws, GH_LOAD_HDR, last_col, ["Assignee", "Metric"])

    for i in range(L.MAX_ASSIGNEES):
        used_row = GH_LOAD_FIRST + 2 * i
        avail_row = used_row + 1
        src = L.GRID_FIRST_DATA_ROW + i

        ws.cell(row=used_row, column=1, value=(
            f'=IF({L.sheet_ref(L.ASSIGNEES)}!A{src}="","",'
            f'{L.sheet_ref(L.ASSIGNEES)}!A{src})')).font = S.FONT_SECTION
        # Labels vanish on unused rows so the pre-built capacity stays invisible
        # until someone actually adds that assignee.
        ws.cell(row=used_row, column=2,
                value=f'=IF($A{used_row}="","","Used")').font = S.FONT_BODY
        ws.cell(row=avail_row, column=2,
                value=f'=IF($A{used_row}="","","Avail")').font = S.FONT_NOTE

        for w in range(L.WEEK_COLS):
            c = GH_WEEK_COL + w
            # Looked up by name rather than by a direct CalcWeek row reference:
            # these rows sit two apart here but one apart there, so a direct
            # reference differs from the row above in R1C1 terms and Excel marks
            # every row after the first with an "inconsistent formula" warning.
            used = ws.cell(row=used_row, column=c, value=(
                f'=IF($A{used_row}="","",IFERROR(INDEX(AwGrid,'
                f'MATCH($A{used_row},AsgNames,0),{w + 1}),0))'))
            avail = ws.cell(row=avail_row, column=c, value=(
                f'=IF($A{used_row}="","",IFERROR(INDEX(CapGrid,'
                f'MATCH($A{used_row},CapNames,0),{w + 1}),0)'
                f'*INDEX(CwFactor,{w + 1}))'))
            for cell in (used, avail):
                cell.number_format = "0.0"
                cell.alignment = S.CENTER
                cell.border = S.BORDER_ALL
            avail.font = S.FONT_NOTE

        # Allocation is capped at capacity, so "used > avail" can never happen.
        # The useful signal is saturation: a week with no slack is a week where
        # this person is the bottleneck holding the plan back.
        span = f"{L.col(GH_WEEK_COL)}{used_row}:{L.col(last_col)}{used_row}"
        ws.conditional_formatting.add(span, Rule(
            type="expression", dxf=DifferentialStyle(fill=S.CF_OVER, font=Font(bold=True)),
            formula=[f'AND({L.col(GH_WEEK_COL)}{used_row}<>"",'
                     f'{L.col(GH_WEEK_COL)}{avail_row}>0,'
                     f'{L.col(GH_WEEK_COL)}{used_row}>='
                     f'{L.col(GH_WEEK_COL)}{avail_row})']))
        ws.conditional_formatting.add(span, Rule(
            type="expression", dxf=DifferentialStyle(fill=S.CF_OK),
            formula=[f'AND(ISNUMBER({L.col(GH_WEEK_COL)}{used_row}),'
                     f'{L.col(GH_WEEK_COL)}{used_row}>0,'
                     f'{L.col(GH_WEEK_COL)}{used_row}<{L.col(GH_WEEK_COL)}{avail_row})']))

        ws.cell(row=used_row, column=1).border = S.BORDER_ALL
        ws.cell(row=avail_row, column=1).border = Border(
            top=Side(style="thin", color="D2D6DC"),
            bottom=Side(style="medium", color=S.INK))

        # How many weeks, from now to the horizon's end, this person has no
        # slack left in. History is excluded — saturation in a week that has
        # already happened is not a lever anyone can pull.
        sat_col = last_col + 1
        first, last_c = L.col(GH_WEEK_COL), L.col(last_col)
        used_range = f"{first}{used_row}:{last_c}{used_row}"
        avail_range = f"{first}{avail_row}:{last_c}{avail_row}"
        flag_range = f"{first}${L.GH_FLAG_ROW}:{last_c}${L.GH_FLAG_ROW}"
        hist_range = f"{first}${L.GH_HIST_ROW}:{last_c}${L.GH_HIST_ROW}"
        ws.cell(row=used_row, column=sat_col, value=(
            f'=IF($A{used_row}="","",SUMPRODUCT(({used_range}>={avail_range})*'
            f'({avail_range}>0)*({flag_range}=1)*({hist_range}=0)))'
        )).alignment = S.CENTER

    last = GH_LOAD_FIRST + 2 * L.MAX_ASSIGNEES - 1
    S.grey_inactive_weeks(ws, GH_WEEK_COL, last_col, GH_LOAD_HDR, last, L.GH_FLAG_ROW)
    ws.cell(row=GH_LOAD_HDR, column=last_col + 1, value="Saturated").font = S.FONT_HDR
    ws.cell(row=GH_LOAD_HDR, column=last_col + 1).fill = S.FILL_OUTPUT_HDR
    return last


def _task_timeline(ws, start_row: int, last_col: int) -> int:
    """One row per task: allocated days per week, shaded where non-zero."""
    S.section(ws, start_row, "Task timeline — allocated days per week", span=last_col)
    hdr = start_row + 1
    _week_header(ws, hdr, last_col, ["Task", "Priority"])
    first = hdr + 1

    for i in range(L.MAX_TASKS):
        r = first + i
        task_row = L.TASK_FIRST_ROW + i
        tw_row = L.CW_TASKWEEK_FIRST + i
        t = L.sheet_ref(L.TASKS)
        ws.cell(row=r, column=1, value=(
            f'=IF({t}!A{task_row}="","",{t}!A{task_row}&"  "&{t}!B{task_row})'))
        ws.cell(row=r, column=2, value=(
            f'=IF({t}!A{task_row}="","",{t}!D{task_row})')).alignment = S.CENTER

        for w in range(L.WEEK_COLS):
            wc = L.col(L.CW_FIRST_WEEK_COL + w)
            cell = ws.cell(row=r, column=GH_WEEK_COL + w, value=(
                f'=IF($A{r}="","",IF({L.sheet_ref(L.CALC_WEEK)}!{wc}{tw_row}=0,"",'
                f'{L.sheet_ref(L.CALC_WEEK)}!{wc}{tw_row}))'))
            cell.number_format = "0.0"
            cell.alignment = S.CENTER
            cell.border = S.BORDER_ALL
        for c in (1, 2):
            ws.cell(row=r, column=c).border = S.BORDER_ALL

    last = first + L.MAX_TASKS - 1
    # ISNUMBER is load-bearing. A week with no work holds "" rather than 0, and
    # Excel ranks any text above any number, so a plain `>0` test is TRUE for an
    # empty cell and paints the whole grid. LibreOffice disagrees and returns
    # FALSE, so this cannot be caught by recalculating there.
    bar = f"{L.col(GH_WEEK_COL)}{first}"
    hist = f"{L.col(GH_WEEK_COL)}${L.GH_HIST_ROW}"
    span = f"{bar}:{L.col(last_col)}{last}"
    # Two mutually exclusive rules rather than one plus an override, so nothing
    # depends on which order Excel happens to evaluate them in.
    ws.conditional_formatting.add(span, Rule(
        type="expression", dxf=DifferentialStyle(fill=S.CF_ACTUAL),
        formula=[f"AND(ISNUMBER({bar}),{bar}>0,{hist}=1)"]))
    ws.conditional_formatting.add(span, Rule(
        type="expression", dxf=DifferentialStyle(fill=S.CF_BAR),
        formula=[f"AND(ISNUMBER({bar}),{bar}>0,{hist}=0)"]))
    S.grey_inactive_weeks(ws, GH_WEEK_COL, last_col, hdr, last, L.GH_FLAG_ROW)
    return last


def _equipment_block(ws, start_row: int, last_col: int) -> int:
    """Two rows per equipment type: units demanded, units available."""
    S.section(ws, start_row, "Equipment — units demanded vs units available",
              span=last_col)
    hdr = start_row + 1
    _week_header(ws, hdr, last_col, ["Equipment", "Metric"])
    first = hdr + 1

    for i in range(L.MAX_EQUIPMENT):
        dem_row = first + 2 * i
        sup_row = dem_row + 1
        src = L.GRID_FIRST_DATA_ROW + i

        ws.cell(row=dem_row, column=1, value=(
            f'=IF({L.sheet_ref(L.EQUIPMENT)}!A{src}="","",'
            f'{L.sheet_ref(L.EQUIPMENT)}!A{src})')).font = S.FONT_SECTION
        ws.cell(row=dem_row, column=2,
                value=f'=IF($A{dem_row}="","","Need")').font = S.FONT_BODY
        ws.cell(row=sup_row, column=2,
                value=f'=IF($A{dem_row}="","","Have")').font = S.FONT_NOTE

        for w in range(L.WEEK_COLS):
            c = GH_WEEK_COL + w
            # By name, for the same reason as the assignee Used rows above.
            need = ws.cell(row=dem_row, column=c, value=(
                f'=IF($A{dem_row}="","",IFERROR(INDEX(EqDemand,'
                f'MATCH($A{dem_row},EqpNames,0),{w + 1}),0))'))
            have = ws.cell(row=sup_row, column=c, value=(
                f'=IF($A{dem_row}="","",IFERROR(INDEX(EqpGrid,'
                f'MATCH($A{dem_row},EqpNames,0),{w + 1}),0))'))
            for cell in (need, have):
                cell.number_format = "0"
                cell.alignment = S.CENTER
                cell.border = S.BORDER_ALL
            have.font = S.FONT_NOTE

        span = f"{L.col(GH_WEEK_COL)}{dem_row}:{L.col(last_col)}{dem_row}"
        ws.conditional_formatting.add(span, Rule(
            type="expression", dxf=DifferentialStyle(fill=S.CF_OVER, font=Font(bold=True)),
            formula=[f'AND({L.col(GH_WEEK_COL)}{dem_row}<>"",'
                     f'{L.col(GH_WEEK_COL)}{dem_row}>{L.col(GH_WEEK_COL)}{sup_row})']))
        ws.cell(row=dem_row, column=1).border = S.BORDER_ALL

    last = first + 2 * L.MAX_EQUIPMENT - 1
    S.grey_inactive_weeks(ws, GH_WEEK_COL, last_col, hdr, last, L.GH_FLAG_ROW)
    return last


def _checks(ws, start_row: int) -> int:
    """Conditions that make a plan wrong, as distinct from merely tight."""
    S.section(ws, start_row, "Checks", span=6)
    S.header_row(ws, start_row + 1, 1, ["Check", "Count"], output=True)

    sub_check = L.abs_range(L.SUBTASKS, L.S_CHECK, L.SUB_FIRST_ROW, L.S_CHECK, N.LAST_SUB_ROW)
    task_check = L.abs_range(L.TASKS, L.T_CHECK, L.TASK_FIRST_ROW, L.T_CHECK, N.LAST_TASK_ROW)

    # Note there is deliberately no "over capacity" check: the spill-over engine
    # caps each week's allocation at what the assignee has, so demand never
    # exceeds supply. Excess work shows up as work that misses the horizon.
    checks = [
        ("Sub-task rows with a problem", f'=COUNTIF({sub_check},"⚠*")'),
        ("Tasks with a problem", f'=COUNTIF({task_check},"⚠*")'),
        ("Assignees with zero capacity all horizon", '=COUNTIFS(AsgNames,"<>",CapTotals,0)'),
        ("Work days that do not fit in the horizon",
         "=ROUND(SUM(SubRemaining)-SUM(CwGrid),2)"),
        ("Assignee-weeks with no slack left",
         "=SUMPRODUCT((AwGrid>=CapGrid)*(CapGrid>0))"),
        ("Equipment-weeks short of units", "=SUMPRODUCT((EqDemand>EqpGrid)*1)"),
    ]
    for i, (label, formula) in enumerate(checks):
        r = start_row + 2 + i
        ws.cell(row=r, column=1, value=label).border = S.BORDER_ALL
        cell = ws.cell(row=r, column=2, value=formula)
        cell.alignment = S.CENTER
        cell.border = S.BORDER_ALL

    span = f"B{start_row + 2}:B{start_row + 1 + len(checks)}"
    ws.conditional_formatting.add(span, Rule(
        type="cellIs", operator="greaterThan", formula=["0"],
        dxf=DifferentialStyle(fill=S.CF_WARN, font=Font(bold=True))))
    ws.conditional_formatting.add(span, Rule(
        type="cellIs", operator="equal", formula=["0"],
        dxf=DifferentialStyle(fill=S.CF_OK)))

    note_row = start_row + 2 + len(checks) + 1
    note = ws.cell(row=note_row, column=1, value=(
        "Saturation and equipment shortage are information, not errors — they show "
        "where the plan is tight. Work that does not fit the horizon is the row to act on."))
    note.font = S.FONT_NOTE
    return note_row


def _plan_health(ws, start_row: int, sat_col: int) -> int:
    """One screen's worth of "is this plan usable, and if not, what would fix
    it" — the quantified Shortfall column on Tasks, rolled up.

    `weeks_to_absorb` assumes the shortfall could be spread across every
    assignee at their average rate, which is optimistic when the real
    constraint is one person or one skill; it is a ballpark, not a promise.
    Bottleneck assignee is read off the Saturated column built alongside the
    assignee-load block above, restricted the same way it is: future weeks,
    inside the horizon.
    """
    S.section(ws, start_row, "Plan health", span=6)

    task_shortfall = L.abs_range(L.TASKS, L.T_SHORTFALL, L.TASK_FIRST_ROW,
                                 L.T_SHORTFALL, N.LAST_TASK_ROW)
    total_shortfall = f"ROUND(SUM({task_shortfall}),2)"
    current_pos = "IFERROR(MATCH(CfgCurrentWeek,CwWeeks,0),1)"
    weeks_left = f"(CfgHorizon-{current_pos}+1)"
    n_assignees = 'MAX(1,COUNTIF(AsgNames,"<>"))'
    # Parenthesised as a whole: it is used as a divisor below, and it is itself
    # a chain of divisions — `total/avg` without the parens would associate as
    # `total/A/B/C` rather than `total/(A/B/C)`.
    avg_weekly_capacity = (
        f"(SUMPRODUCT((CwPos>={current_pos})*(CwPos<=CfgHorizon)*CapGrid*CwFactor)"
        f"/{n_assignees}/MAX(1,{weeks_left}))")

    total_row = start_row + 2
    weeks_row = total_row + 1
    bottleneck_row = weeks_row + 1
    change_row = bottleneck_row + 1

    ws.cell(row=total_row, column=1, value="Total shortfall (days)").border = S.BORDER_ALL
    total_cell = ws.cell(row=total_row, column=2, value=f"={total_shortfall}")
    total_cell.alignment = S.CENTER
    total_cell.border = S.BORDER_ALL

    ws.cell(row=weeks_row, column=1, value="Weeks to absorb it").border = S.BORDER_ALL
    b_total = f"$B{total_row}"
    weeks_cell = ws.cell(row=weeks_row, column=2, value=(
        f'=IF({b_total}<=0,0,IF({avg_weekly_capacity}<=0,"n/a",'
        f'ROUNDUP({b_total}/{avg_weekly_capacity},0)))'))
    weeks_cell.alignment = S.CENTER
    weeks_cell.border = S.BORDER_ALL

    # The Saturated column only has values on a "Used" row, spaced two apart;
    # blanks on the "Avail" rows in between are text, which MAX and MATCH both
    # skip over.
    sat_range = (f"{L.col(sat_col)}{GH_LOAD_FIRST}:"
                 f"{L.col(sat_col)}{GH_LOAD_FIRST + 2 * L.MAX_ASSIGNEES - 1}")
    name_range = f"A{GH_LOAD_FIRST}:A{GH_LOAD_FIRST + 2 * L.MAX_ASSIGNEES - 1}"
    ws.cell(row=bottleneck_row, column=1, value="Bottleneck assignee").border = S.BORDER_ALL
    bottleneck_cell = ws.cell(row=bottleneck_row, column=2, value=(
        f'=IF(MAX({sat_range})<=0,"none",'
        f'INDEX({name_range},MATCH(MAX({sat_range}),{sat_range},0)))'))
    bottleneck_cell.alignment = S.CENTER
    bottleneck_cell.border = S.BORDER_ALL

    ws.cell(row=change_row, column=1, value="Smallest single change").border = S.BORDER_ALL
    b_weeks = f"$B{weeks_row}"
    change_cell = ws.cell(row=change_row, column=2, value=(
        f'=IF({b_total}<=0,"Plan converges — nothing to change.",'
        f'IF(ISNUMBER({b_weeks}),'
        f'"Extend Horizon by "&{b_weeks}&" week(s) would absorb the shortfall '
        f'at average capacity.","No single change sized — check capacity."))'))
    change_cell.alignment = S.CENTER
    change_cell.border = S.BORDER_ALL

    note = ws.cell(row=change_row + 2, column=1, value=(
        "Weeks to absorb assumes the shortfall spreads across everyone at their "
        "average rate — a ballpark, not a promise, if one person or skill is the "
        "real constraint. See Binding constraint on Tasks for the per-task reason."))
    note.font = S.FONT_NOTE
    return change_row + 2


# --- Gantt-Deep ------------------------------------------------------------

GD_WEEK_ROW = 7
GD_DAY_ROW = 8
GD_DATE_ROW = 9
GD_FIRST_ROW = 10
GD_DAY_COL = 5           # day columns start at column E


def build_gantt_deep(ws) -> None:
    S.title(ws, "Gantt — deep dive",
            "Read-only except the two window cells. Day columns, sub-task rows.")

    ws.cell(row=3, column=1, value="Window start week").font = S.FONT_SECTION
    start = ws.cell(row=3, column=2, value="=CfgStartWeek")
    start.number_format = '"WW"0'
    ws.cell(row=4, column=1, value="Weeks to show").font = S.FONT_SECTION
    weeks = ws.cell(row=4, column=2, value=4)
    for cell in (start, weeks):
        cell.border = S.BORDER_ALL
        cell.alignment = S.CENTER
        cell.fill = S.FILL_WARN
    ws.cell(row=4, column=3,
            value=f"Type over either cell. Maximum {L.MAX_DEEP_WEEKS} weeks — "
                  f"wider windows slow recalculation.").font = S.FONT_NOTE

    S.write_flag_row(
        ws, GD_DAY_COL, L.DEEP_DAY_COLS, L.GD_FLAG_ROW,
        lambda i: (f"{L.sheet_ref(L.CALC_DAY)}!"
                   f"{L.col(L.CD_FIRST_DAY_COL + i)}${L.CD_INWINDOW_ROW}"))
    S.write_flag_row(
        ws, GD_DAY_COL, L.DEEP_DAY_COLS, L.GD_HIST_ROW,
        lambda i: (f"IF({L.sheet_ref(L.CALC_DAY)}!"
                   f"{L.col(L.CD_FIRST_DAY_COL + i)}${L.CD_POS_ROW}"
                   f"<IFERROR(MATCH(CfgCurrentWeek,CwWeeks,0),1),1,0)"))
    _deep_headers(ws)
    _deep_rows(ws)

    S.widths(ws, {"A": 12, "B": 12, "C": 26, "D": 14})
    for i in range(L.DEEP_DAY_COLS):
        ws.column_dimensions[L.col(GD_DAY_COL + i)].width = 5.5
    ws.freeze_panes = L.col(GD_DAY_COL) + str(GD_FIRST_ROW)


def _deep_headers(ws) -> None:
    S.header_row(ws, GD_WEEK_ROW, 1,
                 ["Parent", "Sub-task", "Name", "Assignee"], output=True)
    for row in (GD_DAY_ROW, GD_DATE_ROW):
        for c in range(1, GD_DAY_COL):
            ws.cell(row=row, column=c).fill = S.FILL_OUTPUT_HDR

    for i in range(L.DEEP_DAY_COLS):
        c = GD_DAY_COL + i
        src = L.col(L.CD_FIRST_DAY_COL + i)
        cd = L.sheet_ref(L.CALC_DAY)

        week = ws.cell(row=GD_WEEK_ROW, column=c, value=f"={cd}!{src}${L.CD_LABEL_ROW}")
        day = ws.cell(row=GD_DAY_ROW, column=c,
                      value=f'="{L.DAY_NAMES[i % L.WORKDAYS_PER_WEEK]}"')
        date = ws.cell(row=GD_DATE_ROW, column=c, value=f"={cd}!{src}${L.CD_DATE_ROW}")
        date.number_format = "dd/mm"

        for cell in (week, day, date):
            cell.fill = S.FILL_OUTPUT_HDR
            cell.font = S.FONT_HDR
            cell.alignment = S.CENTER
            cell.border = S.BORDER_ALL

    last = GD_DAY_COL + L.DEEP_DAY_COLS - 1
    cd = L.sheet_ref(L.CALC_DAY)
    # Grey out columns outside the window; they cannot be hidden dynamically
    # without macros. Read from this sheet's own hidden flag row, because a rule
    # that reaches into CalcDay would be promoted to an x14 extension that
    # openpyxl cannot preserve.
    for row in (GD_WEEK_ROW, GD_DAY_ROW, GD_DATE_ROW):
        ws.conditional_formatting.add(
            f"{L.col(GD_DAY_COL)}{row}:{L.col(last)}{row}",
            Rule(type="expression", dxf=DifferentialStyle(fill=S.CF_WEEKEND,
                                                          font=Font(color=S.cf_color(S.MUTED))),
                 formula=[f"{L.col(GD_DAY_COL)}${L.GD_FLAG_ROW}=0"]))


def _deep_rows(ws) -> None:
    cd = L.sheet_ref(L.CALC_DAY)
    last_col = GD_DAY_COL + L.DEEP_DAY_COLS - 1

    for i in range(L.MAX_SUBTASKS):
        r = GD_FIRST_ROW + i
        cdr = L.CD_FIRST_ROW + i

        ws.cell(row=r, column=1, value=f'=IF({cd}!$F{cdr}="","",{cd}!$F{cdr})')
        ws.cell(row=r, column=2, value=f'=IF({cd}!$B{cdr}="","",{cd}!$B{cdr})')
        ws.cell(row=r, column=3, value=(
            f'=IF({cd}!$B{cdr}="","",IFERROR(INDEX(SubName,'
            f'MATCH({cd}!$A{cdr},SubRank,0)),""))'))
        ws.cell(row=r, column=4, value=f'=IF({cd}!$C{cdr}="","",{cd}!$C{cdr})')

        for j in range(L.DEEP_DAY_COLS):
            src = L.col(L.CD_FIRST_DAY_COL + j)
            cell = ws.cell(row=r, column=GD_DAY_COL + j, value=(
                f'=IF({cd}!$B{cdr}="","",IF({cd}!{src}{cdr}=0,"",{cd}!{src}{cdr}))'))
            cell.number_format = "0.0"
            cell.alignment = S.CENTER
            cell.border = S.BORDER_ALL
        for c in range(1, GD_DAY_COL):
            ws.cell(row=r, column=c).border = S.BORDER_ALL

    last_row = GD_FIRST_ROW + L.MAX_SUBTASKS - 1
    body = f"{L.col(GD_DAY_COL)}{GD_FIRST_ROW}:{L.col(last_col)}{last_row}"
    # Same reason as the task timeline: an idle day holds "" and Excel treats
    # text as greater than any number.
    bar = f"{L.col(GD_DAY_COL)}{GD_FIRST_ROW}"
    hist = f"{L.col(GD_DAY_COL)}${L.GD_HIST_ROW}"
    ws.conditional_formatting.add(body, Rule(
        type="expression", dxf=DifferentialStyle(fill=S.CF_ACTUAL),
        formula=[f"AND(ISNUMBER({bar}),{bar}>0,{hist}=1)"]))
    ws.conditional_formatting.add(body, Rule(
        type="expression", dxf=DifferentialStyle(fill=S.CF_BAR),
        formula=[f"AND(ISNUMBER({bar}),{bar}>0,{hist}=0)"]))
    ws.conditional_formatting.add(body, Rule(
        type="expression", dxf=DifferentialStyle(fill=S.CF_WEEKEND),
        formula=[f"{L.col(GD_DAY_COL)}${L.GD_FLAG_ROW}=0"]))

    # Same hide-the-repeats treatment as the Sub-Tasks tab.
    ws.conditional_formatting.add(
        f"A{GD_FIRST_ROW}:A{last_row}",
        Rule(type="expression", dxf=DifferentialStyle(font=Font(color=S.cf_color("FFFFFF"))),
             formula=[f'AND($A{GD_FIRST_ROW}<>"",$A{GD_FIRST_ROW}=$A{GD_FIRST_ROW - 1})']))
    ws.conditional_formatting.add(
        f"A{GD_FIRST_ROW}:{L.col(last_col)}{last_row}",
        Rule(type="expression",
             dxf=DifferentialStyle(border=Border(top=Side(style="medium", color=S.cf_color(S.INK)))),
             formula=[f'AND($A{GD_FIRST_ROW}<>"",$A{GD_FIRST_ROW}<>$A{GD_FIRST_ROW - 1})']))
