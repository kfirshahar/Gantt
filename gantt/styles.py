"""Shared fills, fonts and helpers so every sheet reads as one document."""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Palette
INK = "1F2933"
MUTED = "7B8794"
BAND_INPUT = "1F4E79"       # header band on tabs the user edits
BAND_OUTPUT = "3B5A40"      # header band on read-only views
BAND_SECTION = "E4E7EB"     # block sub-headings

OVER = "F8C9C9"
WARN = "FDE9C9"
OK = "D8EFD8"
BAR = "9FC5E8"
WEEKEND = "EEF1F4"
ACTUAL = "CBD5DC"      # history: settled, deliberately duller than the plan

FILL_INPUT_HDR = PatternFill("solid", fgColor=BAND_INPUT)
FILL_OUTPUT_HDR = PatternFill("solid", fgColor=BAND_OUTPUT)
FILL_SECTION = PatternFill("solid", fgColor=BAND_SECTION)
FILL_DERIVED = PatternFill("solid", fgColor="F5F7FA")
FILL_OVER = PatternFill("solid", fgColor=OVER)
FILL_WARN = PatternFill("solid", fgColor=WARN)
FILL_OK = PatternFill("solid", fgColor=OK)
FILL_BAR = PatternFill("solid", fgColor=BAR)
FILL_WEEKEND = PatternFill("solid", fgColor=WEEKEND)


def cf_fill(rgb: str) -> PatternFill:
    """A fill for use inside a conditional-format DifferentialStyle.

    Two things differ from an ordinary cell fill and Excel needs both. A dxf
    solid fill takes its colour from bgColor, not fgColor — Excel writes its own
    that way and ignores fgColor here. And the colour needs an explicit opaque
    alpha, because openpyxl otherwise pads a bare six-digit value to `00RRGGBB`,
    which is fully transparent. LibreOffice forgives both; Excel renders nothing.
    """
    return PatternFill(patternType="solid", bgColor="FF" + rgb.lstrip("#")[-6:])


def cf_color(rgb: str) -> str:
    """An opaque ARGB colour for a conditional-format font or border.

    Same alpha trap as `cf_fill`: a bare six-digit value is padded to `00RRGGBB`.
    """
    return "FF" + rgb.lstrip("#")[-6:]


CF_OVER = cf_fill(OVER)
CF_WARN = cf_fill(WARN)
CF_OK = cf_fill(OK)
CF_BAR = cf_fill(BAR)
CF_WEEKEND = cf_fill(WEEKEND)
CF_ACTUAL = cf_fill(ACTUAL)

FONT_TITLE = Font(bold=True, size=14, color=INK)
FONT_HDR = Font(bold=True, size=10, color="FFFFFF")
FONT_SECTION = Font(bold=True, size=10, color=INK)
FONT_DERIVED = Font(italic=True, size=10, color=MUTED)
FONT_NOTE = Font(italic=True, size=9, color=MUTED)
FONT_BODY = Font(size=10, color=INK)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

THIN = Side(style="thin", color="D2D6DC")
THICK = Side(style="medium", color=INK)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_GROUP_TOP = Border(left=THIN, right=THIN, top=THICK, bottom=THIN)


def title(ws, text: str, note: str | None = None) -> None:
    ws["A1"] = text
    ws["A1"].font = FONT_TITLE
    if note:
        ws["A2"] = note
        ws["A2"].font = FONT_NOTE


def header_row(ws, row: int, first_col: int, values, output: bool = False) -> None:
    """Write a banded header row starting at `first_col`."""
    fill = FILL_OUTPUT_HDR if output else FILL_INPUT_HDR
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=first_col + offset, value=value)
        cell.fill = fill
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = BORDER_ALL


def section(ws, row: int, text: str, span: int = 12) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SECTION
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION


def widths(ws, mapping: dict) -> None:
    for letter, width in mapping.items():
        ws.column_dimensions[letter].width = width


def mark_derived(cell) -> None:
    cell.font = FONT_DERIVED
    cell.fill = FILL_DERIVED


def write_flag_row(ws, first_col: int, count: int, row: int, source) -> None:
    """Mirror a flag from another sheet into a hidden row on this one.

    Conditional formatting has to read the flag from its own sheet, but nothing
    stops an ordinary cell formula fetching it. `source` is called with the
    0-based column offset and returns the formula body. Hidden so it never shows.
    """
    for i in range(count):
        ws.cell(row=row, column=first_col + i, value=f"={source(i)}")
    ws.row_dimensions[row].hidden = True


def grey_inactive_weeks(ws, first_col: int, last_col: int,
                        first_row: int, last_row: int, flag_row: int) -> None:
    """Dim week columns that fall outside the horizon set on Config.

    The columns are always present; this is what makes them read as switched
    off rather than as weeks with no work in them. `flag_row` must be on this
    same sheet — see `write_flag_row`.
    """
    from openpyxl.formatting.rule import Rule
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.utils import get_column_letter

    first = get_column_letter(first_col)
    ws.conditional_formatting.add(
        f"{first}{first_row}:{get_column_letter(last_col)}{last_row}",
        Rule(type="expression",
             dxf=DifferentialStyle(fill=CF_WEEKEND, font=Font(color=cf_color(MUTED))),
             formula=[f"{first}${flag_row}=0"]))
