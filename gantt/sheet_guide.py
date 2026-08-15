"""The Guide tab.

Kept inside the workbook rather than alongside it, so the instructions travel
with the file when it is copied for another project.
"""

from openpyxl.styles import Alignment, Font

from . import layout as L, styles as S

BODY_COL = 2
WRAP = Alignment(wrap_text=True, vertical="top")

# (heading, [line, ...]). A line beginning with a digit renders as a numbered
# step; everything else as a plain paragraph.
SECTIONS: list[tuple[str, list[str]]] = [
    ("Colour tells you what you may touch", [
        "White cells — you type here.",
        "Grey italic — computed. Typing in one replaces a formula and breaks the "
        "schedule from that row down. If you do it by accident, press Ctrl+Z.",
        "Amber — the few settings that change what every other tab shows: "
        "Horizon on Config, and the two window cells on Gantt-Deep.",
        "Green header — a read-only output tab.",
    ]),
    ("Fill the tabs in this order (first time only)", [
        "1. Config — project year, start week, horizon. The Complexity table is the "
        "dial that turns the word 'Complex' into a number of days.",
        "2. Assignees — names and proficiency. 1.00 is an average performer; 1.20 "
        "finishes a job in 1/1.2 of the base days; 0.80 takes longer.",
        "3. Capacity — how many days each person has in a normal five-day week, net "
        "of vacation and part-time hours. Do not subtract company holidays here; "
        "the Holidays tab does that for you.",
        "4. Equipment — how many units of each type exist in the shared pool each week.",
        "5. Holidays — company-wide non-working days. Each one shortens the week for "
        "everybody and cuts that week's capacity in proportion: one holiday makes the "
        "week worth 4/5 of what you entered on Capacity.",
        "6. Tasks — one row per deliverable.",
        "7. Sub-Tasks — the actual work. All effort comes from here.",
    ]),
    ("Recording what actually happened", [
        "Actual start WW and Actual end WW on Sub-Tasks are where you say when the "
        "work really ran. Leave the end blank while it is still going.",
        "The Gantt then draws history in the weeks behind the current one, in a "
        "duller colour than the plan, so you can see what was done alongside what "
        "is still to come. Effort already spent is spread evenly across the "
        "recorded span — it is history, not a measurement.",
        "Nothing about history feeds back into scheduling. It sets how much work "
        "is left and nothing else, which is what keeps the two Gantt views telling "
        "the same story.",
        "A past week can therefore show more days used than were available. That "
        "is not an error: it means the work really did overrun what Capacity said, "
        "and hiding it would be the wrong answer.",
    ]),
    ("Where the numbers on Gantt-High come from", [
        "Avail is not simply what you typed on Capacity. It is that figure reduced "
        "pro-rata for company holidays, so a week containing one holiday shows 4/5 of "
        "it. This is the only place holidays change a total.",
        "Because of that, a person is never given more than one day of work in a day: "
        "their daily rate is just their weekly figure divided by five, and a holiday "
        "removes the day rather than squeezing the same work into the days that remain.",
    ]),
    ("Where 'now' is, and why the horizon does not move", [
        "Config has a Current week. Weeks before it are history; planning starts "
        "there. Move it on as the project progresses — the horizon itself stays "
        "anchored at the project start.",
        "So an Earliest start WW that has already gone by is not an error. It "
        "simply means the work should start as soon as possible, and it is "
        "scheduled from the current week instead.",
        "'Starts beyond the horizon' now means only what it says: a start week "
        "later than the last week the plan covers.",
    ]),
    ("How the scheduling works", [
        "You never type a duration. Each sub-task's effort is the base days for its "
        "complexity divided by its assignee's proficiency. A task's effort is the sum "
        "of its sub-tasks.",
        "What actually competes for capacity is Remaining, not Effort: Status TODO or "
        "In Progress claims effort minus % done, and Status Done claims nothing at all "
        "regardless of what % done says. Mark work Done and it stops occupying anyone's "
        "week.",
        "Sub-tasks are then placed into weeks in a fixed queue order. Each one takes "
        "whatever days its assignee has left in that week, and spills into the next "
        "week, and the next, until its remaining effort is used up.",
        "So duration is an outcome of capacity, not something you set. Give someone "
        "fewer days in a week and the work simply takes longer.",
        "A task's own Status and Remaining are never typed in — they are rolled up "
        "from its sub-tasks, the same way its Effort is. Done means every sub-task is "
        "Done; TODO means none has started; anything else is In Progress.",
    ]),
    ("Rank — you never type it, and it cannot be broken", [
        "Rank is only the position in that queue: rank 1 is served first, rank 2 next. "
        "It is recalculated from scratch every time anything changes.",
        "It is worked out from, in order: the parent task's Priority, then the parent's "
        "Earliest start WW, then the order of rows on Tasks, then the sub-task's "
        "position under its parent.",
        "It exists because two sub-tasks may want the same person in the same week, and "
        "something has to decide who gets the days. Rank is that decision, and it is why "
        "P1 work is never pushed aside by P2 work.",
        "You steer it with exactly two columns on Tasks: Priority and Earliest start WW. "
        "Nothing else affects it. Changing a priority reshuffles the whole queue and can "
        "move tasks you did not touch — that is correct, because everyone draws from the "
        "same pool of days.",
    ]),
    ("Adding a new task", [
        "1. On Tasks, use the next empty row. Fill only columns A to H: ID, name, "
        "category, priority, complexity, equipment, default assignee, earliest start WW.",
        "2. On Sub-Tasks, add one row per piece of work. Fill only Parent task, Sub-task "
        "name and Complexity. Leave Assignee blank to inherit the parent's default — "
        "fill it in only for the exceptions. Status defaults to TODO and % done to 0; "
        "update them as work happens.",
        "3. Confirm the Check column reads 'ok' on both tabs.",
        "The computed columns are already filled in down to row 601 on Sub-Tasks and "
        "row 31 on Tasks. Do not copy formulas down — just type in the white cells.",
    ]),
    ("What the Check messages mean", [
        "done (Tasks) — every sub-task is finished. Not a warning; it is there so "
        "completed work stops being reported as a problem.",
        "in progress (Tasks) — under way, with nothing left to schedule.",
        "no complexity — pick one from the dropdown. This is the usual one on a row you "
        "have just added.",
        "no assignee — neither this sub-task nor its parent task names anybody. Set the "
        "parent's Default assignee, or fill in Assignee on this row.",
        "unknown parent — the Parent task does not exist on Tasks. This only happens if "
        "you paste a value instead of choosing from the dropdown.",
        "check proficiency — that assignee's proficiency is blank or zero, so effort "
        "cannot be divided by it.",
        "no actual start — the sub-task is under way or finished but never says "
        "when it began, so there is nothing to draw as history.",
        "no sub-tasks (Tasks) — the task has no work underneath it, so its effort is zero.",
        "start week not in the grid (Tasks) — the Earliest start WW names no week the "
        "plan covers at all. Check the number.",
        "starts beyond the horizon (Tasks) — the start is later than the last week "
        "shown. Widen Horizon on Config, or bring the task forward.",
        "not scheduled (Tasks) — it has effort but nothing landed anywhere. Usually the "
        "assignee has no capacity at all.",
        "overruns horizon (Tasks) — the work does not finish inside the horizon. Widen "
        "Horizon, add capacity, or cut scope.",
    ]),
    ("Reading the Gantt tabs", [
        "Gantt-High — red on an assignee's Used row means saturated: no slack left that "
        "week, so that person is the constraint. Green means there is room. Red on an "
        "equipment Need row means more units are wanted than exist.",
        "Nobody is ever given more days than they have, so a week can never be "
        "over-allocated. Too much work shows up instead as 'Work days that do not fit "
        "in the horizon' in the Checks block.",
        "Gantt-Deep — a scrolling window of up to 8 weeks at day level. Change 'Window "
        "start week' to move it; it is not limited to the start of the project.",
        "Greyed-out week columns are outside the Horizon set on Config. They are not "
        "empty weeks, they are switched off.",
    ]),
    ("Things that will bite", [
        "Sub-task IDs are positional: T-01.01 is simply the first T-01 row counting from "
        "the top. Sorting the tab renumbers them. Filtering is safe; sorting is not "
        "recommended.",
        "Keep all the rows of one parent task together.",
        "Earliest start WW is a real calendar week. If the plan runs into next year, "
        "type 2 to mean WW02 of next year — the headers show the year for you.",
        "Built-in limits: 30 tasks, 600 sub-tasks, 10 assignees, 10 equipment types, "
        "26 weeks. Horizon on Config can be anything up to 26.",
        "CalcWeek and CalcDay are hidden working sheets. You never need them, but "
        "right-click a tab and choose Unhide if you want to look.",
    ]),
]


def build_guide(ws) -> None:
    S.title(ws, "How to use this file",
            "Read the first two sections; the rest is reference.")

    row = 4
    for heading, lines in SECTIONS:
        cell = ws.cell(row=row, column=1, value=heading)
        cell.font = Font(bold=True, size=11, color=S.INK)
        cell.fill = S.FILL_SECTION
        ws.cell(row=row, column=BODY_COL).fill = S.FILL_SECTION
        row += 1

        for line in lines:
            if line[0].isdigit():
                marker, _, text = line.partition(". ")
                num = ws.cell(row=row, column=1, value=marker + ".")
                num.font = S.FONT_SECTION
                # Must match the body's vertical alignment, or the number floats
                # to the bottom of a tall wrapped row and reads as belonging to
                # the line beneath it.
                num.alignment = WRAP
            else:
                text = line
            body = ws.cell(row=row, column=BODY_COL, value=text)
            body.font = S.FONT_BODY
            body.alignment = WRAP
            # Row height is left unset so Excel auto-fits the wrapped text;
            # estimating it here produced ragged gaps.
            row += 1
        row += 1

    S.widths(ws, {"A": 4, "B": 105})
    ws.sheet_view.showGridLines = False
