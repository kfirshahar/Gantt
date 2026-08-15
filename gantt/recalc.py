"""Recalculating a generated workbook so its formula results can be checked.

openpyxl writes formulas but never evaluates them, so verifying the numbers
needs a real spreadsheet engine. Two are supported, and which one runs matters
more than it might appear.

LibreOffice is available everywhere and recalculates numbers faithfully, but it
does not reproduce Excel's formatting or comparison semantics. Every rendering
bug this project has shipped — a differential fill keyed on fgColor rather than
bgColor, a colour padded to a transparent alpha, and `"" > 0` evaluating true —
was invisible to it. Excel, driven over COM, is the engine the file is actually
opened in, so on Windows it is the more trustworthy of the two and is preferred.

Selection, in order:

- ``GANTT_RECALC`` — ``excel``, ``libreoffice`` or ``none`` to force a choice.
- otherwise Excel if this is Windows and pywin32 is importable,
- otherwise LibreOffice if it can be found,
- otherwise nothing, and callers skip their numeric checks.

``GANTT_SOFFICE`` overrides the search for the LibreOffice binary.
"""

from __future__ import annotations

import os
import platform
import shutil
import subprocess
from pathlib import Path

ENV_BACKEND = "GANTT_RECALC"
ENV_SOFFICE = "GANTT_SOFFICE"

_SOFFICE_CANDIDATES = {
    "Darwin": [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ],
    "Windows": [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ],
    "Linux": [
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/snap/bin/libreoffice",
    ],
}


def find_libreoffice() -> Path | None:
    """The soffice binary, or None. An explicit override wins outright."""
    explicit = os.environ.get(ENV_SOFFICE)
    if explicit:
        path = Path(explicit)
        return path if path.exists() else None

    for candidate in _SOFFICE_CANDIDATES.get(platform.system(), []):
        if Path(candidate).exists():
            return Path(candidate)

    for name in ("soffice", "soffice.exe", "libreoffice"):
        found = shutil.which(name)
        if found:
            return Path(found)
    return None


def excel_available() -> bool:
    """Whether Excel can be driven over COM from this interpreter."""
    if platform.system() != "Windows":
        return False
    try:
        import win32com.client  # noqa: F401
    except ImportError:
        return False
    return True


class Backend:
    """Recalculates a workbook and returns a path to the recalculated copy."""

    name = "none"

    def recalculate(self, src: Path, outdir: Path) -> Path:
        raise NotImplementedError


class LibreOffice(Backend):
    name = "libreoffice"

    def __init__(self, soffice: Path) -> None:
        self.soffice = soffice

    def recalculate(self, src: Path, outdir: Path) -> Path:
        outdir.mkdir(parents=True, exist_ok=True)
        subprocess.run(
            [str(self.soffice), "--headless", "--norestore",
             "--convert-to", "xlsx", "--outdir", str(outdir), str(src)],
            check=True, capture_output=True, timeout=300)
        return outdir / src.name


class ExcelCOM(Backend):
    """Drives the installed Excel. Windows only.

    `DispatchEx` deliberately starts a private instance so the run cannot
    disturb, or be disturbed by, a copy the user already has open.

    COM is initialised and torn down explicitly. pytest runs tests on a worker
    thread, and a thread that touches COM without initialising it first is the
    usual source of the "Windows fatal exception" faulthandler prints during
    interpreter shutdown. References are dropped in creation order before Quit
    for the same reason: a live workbook reference keeps Excel alive and the
    process lingers after the run.
    """

    name = "excel"

    def recalculate(self, src: Path, outdir: Path) -> Path:
        import pythoncom
        import win32com.client

        outdir.mkdir(parents=True, exist_ok=True)
        dst = outdir / src.name
        if dst.exists():
            dst.unlink()          # SaveAs prompts rather than overwriting
        xlOpenXMLWorkbook = 51

        pythoncom.CoInitialize()
        app = book = None
        try:
            app = win32com.client.DispatchEx("Excel.Application")
            app.Visible = False
            app.DisplayAlerts = False
            app.AskToUpdateLinks = False
            book = app.Workbooks.Open(str(src.resolve()))
            app.CalculateFullRebuild()
            book.SaveAs(str(dst.resolve()), FileFormat=xlOpenXMLWorkbook)
            book.Close(SaveChanges=False)
            book = None
        finally:
            if book is not None:
                try:
                    book.Close(SaveChanges=False)
                except Exception:
                    pass
            if app is not None:
                try:
                    app.Quit()
                except Exception:
                    pass
            del book, app
            pythoncom.CoUninitialize()
        return dst


def get_backend() -> Backend | None:
    """The backend to use, honouring GANTT_RECALC, or None if none is usable."""
    choice = os.environ.get(ENV_BACKEND, "auto").strip().lower()

    if choice == "none":
        return None

    if choice in ("excel", "auto") and excel_available():
        return ExcelCOM()
    if choice == "excel":
        return None

    if choice in ("libreoffice", "auto"):
        soffice = find_libreoffice()
        if soffice:
            return LibreOffice(soffice)
    return None


def describe() -> str:
    backend = get_backend()
    if backend is None:
        return ("no recalculation backend found; set GANTT_SOFFICE to a soffice "
                "binary, or install pywin32 on Windows to drive Excel")
    if isinstance(backend, LibreOffice):
        return f"libreoffice at {backend.soffice}"
    return "excel via COM"


def fingerprint(path: Path) -> str:
    """Which application last wrote this workbook.

    Every engine stamps its own name into docProps/app.xml on save, so this
    reports what actually recalculated a file rather than what was asked to.
    openpyxl writes "Microsoft Excel Compatible / Openpyxl", LibreOffice writes
    its name and build, and Excel writes "Microsoft Excel".
    """
    import re
    import zipfile

    with zipfile.ZipFile(path) as archive:
        if "docProps/app.xml" not in archive.namelist():
            return "(no application recorded)"
        app = archive.read("docProps/app.xml").decode("utf-8", "replace")
    found = re.search(r"<Application>(.*?)</Application>", app)
    return found.group(1) if found else "(no application recorded)"


def verify(workdir: Path | None = None) -> dict:
    """End-to-end proof that the selected engine really recalculates.

    Builds a throwaway workbook holding a formula with no cached result, runs it
    through the backend, then reports both the value that came back and the
    application that produced it. A correct value proves something evaluated the
    formula; the fingerprint proves which something.
    """
    import tempfile

    from openpyxl import Workbook, load_workbook

    backend = get_backend()
    result = {"backend": backend.name if backend else None,
              "selected": describe(), "computed": None,
              "fingerprint": None, "ok": False}
    if backend is None:
        return result

    created = workdir is None
    base = Path(tempfile.mkdtemp(prefix="gantt-verify-")) if created else Path(workdir)
    try:
        src = base / "probe.xlsx"
        book = Workbook()
        book.active["A1"] = "=1+1"      # openpyxl caches nothing
        book.save(src)

        out = backend.recalculate(src, base / "out")
        result["computed"] = load_workbook(out, data_only=True).active["A1"].value
        result["fingerprint"] = fingerprint(out)
        result["ok"] = result["computed"] == 2
    finally:
        if created:
            import shutil as _shutil
            _shutil.rmtree(base, ignore_errors=True)
    return result


def _main() -> int:
    report = verify()
    print(f"selected     : {report['selected']}")
    print(f"backend      : {report['backend'] or '(none)'}")
    print(f"computed 1+1 : {report['computed']!r}  (expected 2)")
    print(f"written by   : {report['fingerprint']}")

    if not report["ok"]:
        print("\nFAIL: nothing recalculated the probe formula.")
        return 1

    engine = (report["fingerprint"] or "").lower()
    if report["backend"] == "excel" and "microsoft excel" not in engine:
        print("\nFAIL: the excel backend was selected but the file was not "
              "written by Excel.")
        return 1
    if report["backend"] == "libreoffice" and "libreoffice" not in engine:
        print("\nFAIL: the libreoffice backend was selected but the file was "
              "not written by LibreOffice.")
        return 1

    print("\nOK: the selected engine really recalculated the workbook.")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
