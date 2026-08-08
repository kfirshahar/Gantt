"""Structural and numerical verification of the generated template.

The numerical half depends on LibreOffice: openpyxl cannot evaluate formulas,
so the workbook is recalculated headlessly and the resulting values are compared
against `gantt.reference`, an independent Python model of the same algorithm.
"""

import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from gantt import calendar_utils as C, demo, layout as L, names as N, reference  # noqa: E402
from gantt.build import build  # noqa: E402

SOFFICE = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
TOL = 1e-6


@pytest.fixture(scope="session")
def built(tmp_path_factory) -> Path:
    out = tmp_path_factory.mktemp("build") / "Gantt_Template.xlsx"
    return build(out)


@pytest.fixture(scope="session")
def formulas(built):
    return load_workbook(built, data_only=False)


@pytest.fixture(scope="session")
def values(built, tmp_path_factory):
    """The workbook after LibreOffice has recalculated every formula."""
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed; numerical checks skipped")
    outdir = tmp_path_factory.mktemp("recalc")
    subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(built)],
        check=True, capture_output=True, timeout=300)
    return load_workbook(outdir / built.name, data_only=True)


@pytest.fixture(scope="session")
def solved():
    return reference.solve(window_start=demo.START_WEEK, window_weeks=4)


# --- Structure -------------------------------------------------------------

def test_all_sheets_present(formulas):
    assert set(formulas.sheetnames) == set(
        L.INPUT_SHEETS + L.OUTPUT_SHEETS + L.HIDDEN_SHEETS)


def test_calc_sheets_hidden(formulas):
    for name in L.HIDDEN_SHEETS:
        assert formulas[name].sheet_state == "hidden"


def test_defined_names_registered(formulas):
    for name in N.NAMES:
        assert name in formulas.defined_names, f"{name} missing"


def test_dropdowns_on_cross_referencing_columns(formulas):
    expected = {
        L.TASKS: {"PrioNames", "CplxNames", "EqpNames", "AsgNames"},
        L.SUBTASKS: {"TaskIDs", "CplxNames", "AsgNames"},
    }
    for sheet, names in expected.items():
        found = {dv.formula1.lstrip("=") for dv in formulas[sheet].data_validations.dataValidation}
        assert names <= found, f"{sheet} missing {names - found}"


def test_parent_column_is_not_merged(formulas):
    """A real merge would break sort and AutoFilter and empty every repeat cell."""
    ws = formulas[L.SUBTASKS]
    assert not ws.merged_cells.ranges
    # every demo row still carries its real parent id
    for i, (parent, *_rest) in enumerate(demo.subtasks()):
        assert ws.cell(row=L.SUB_FIRST_ROW + i, column=L.S_PARENT).value == parent


def test_no_formula_references_a_stale_sheet(formulas):
    known = set(formulas.sheetnames)
    for ws in formulas.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and "!" in v:
                    for chunk in v.replace("'", "").split("!")[:-1]:
                        name = chunk.split("(")[-1].split(",")[-1].split("+")[-1].strip()
                        if name and name[0].isalpha() and " " not in name and "-" in name:
                            assert name in known or name in N.NAMES, f"{ws.title}: {v}"


# --- Numbers ---------------------------------------------------------------

def _week_cols():
    return {demo.START_WEEK + i: L.CW_FIRST_WEEK_COL + i for i in range(demo.HORIZON)}


def test_effort_matches_reference(values, solved):
    ws = values[L.SUBTASKS]
    by_id = {f"{s.parent}.{s.index:02d}": s for s in solved}
    for r in range(L.SUB_FIRST_ROW, L.SUB_FIRST_ROW + len(solved)):
        sub_id = ws.cell(row=r, column=L.S_ID).value
        got = ws.cell(row=r, column=L.S_EFFORT).value
        assert abs(got - by_id[sub_id].effort) < TOL, sub_id


def test_ranks_are_unique_and_match_reference(values, solved):
    ws = values[L.SUBTASKS]
    by_id = {f"{s.parent}.{s.index:02d}": s for s in solved}
    ranks = []
    for r in range(L.SUB_FIRST_ROW, L.SUB_FIRST_ROW + len(solved)):
        sub_id = ws.cell(row=r, column=L.S_ID).value
        rank = ws.cell(row=r, column=L.S_RANK).value
        ranks.append(rank)
        assert rank == by_id[sub_id].rank, sub_id
    assert len(set(ranks)) == len(ranks), "ranks must be unique or INDEX/MATCH breaks"


def test_weekly_allocations_match_reference(values, solved):
    ws = values[L.CALC_WEEK]
    cols = _week_cols()
    by_rank = {s.rank: s for s in solved}
    for i in range(len(solved)):
        row = L.CW_FIRST_ROW + i
        rank = ws.cell(row=row, column=L.CW_RANK).value
        expected = by_rank[rank]
        assert ws.cell(row=row, column=L.CW_SUBID).value == f"{expected.parent}.{expected.index:02d}"
        for week, col in cols.items():
            got = ws.cell(row=row, column=col).value
            assert abs(got - expected.weekly[week]) < TOL, f"rank {rank} {week}"


def test_no_week_exceeds_assignee_capacity(values, solved):
    load = reference.assignee_load(solved)
    weeks = list(_week_cols())
    for name, per_week in load.items():
        for i, week in enumerate(weeks):
            cap = demo.CAPACITY[name][i]
            assert per_week[week] <= cap + TOL, f"{name} {week}"


def test_assignee_load_block_matches_reference(values, solved):
    ws = values[L.CALC_WEEK]
    load = reference.assignee_load(solved)
    cols = _week_cols()
    for i, (name, _) in enumerate(demo.ASSIGNEES):
        row = L.CW_AWWEEK_FIRST + i
        assert ws.cell(row=row, column=1).value == name
        for week, col in cols.items():
            assert abs(ws.cell(row=row, column=col).value - load[name][week]) < TOL


def test_task_rollup_matches_reference(values, solved):
    ws = values[L.CALC_WEEK]
    loads = reference.task_load(solved)
    cols = _week_cols()
    for i, task in enumerate(demo.TASKS):
        row = L.CW_TASKWEEK_FIRST + i
        assert ws.cell(row=row, column=1).value == task[0]
        for week, col in cols.items():
            assert abs(ws.cell(row=row, column=col).value - loads[task[0]][week]) < TOL


def test_equipment_demand_matches_reference(values, solved):
    ws = values[L.CALC_WEEK]
    demand = reference.equipment_demand(solved)
    cols = _week_cols()
    for i, name in enumerate(demo.EQUIPMENT):
        row = L.CW_EQPWEEK_FIRST + i
        assert ws.cell(row=row, column=1).value == name
        for week, col in cols.items():
            assert ws.cell(row=row, column=col).value == demand[name][week], f"{name} {week}"


def test_task_effort_is_sum_of_subtasks(values, solved):
    ws = values[L.TASKS]
    totals: dict[str, float] = {}
    for s in solved:
        totals[s.parent] = totals.get(s.parent, 0.0) + s.effort
    for i, task in enumerate(demo.TASKS):
        row = L.TASK_FIRST_ROW + i
        assert ws.cell(row=row, column=L.T_N_SUBS).value == task[8]
        assert abs(ws.cell(row=row, column=L.T_EFFORT).value - round(totals[task[0]], 2)) < 0.01


def test_daily_allocations_match_reference(values, solved):
    """The deep-dive must reproduce the same schedule at day granularity."""
    ws = values[L.CALC_DAY]
    by_rank = {s.rank: s for s in solved}
    dates = {}
    for j in range(L.DEEP_DAY_COLS):
        col = L.CD_FIRST_DAY_COL + j
        d = ws.cell(row=L.CD_DATE_ROW, column=col).value
        in_window = ws.cell(row=L.CD_INWINDOW_ROW, column=col).value
        if in_window:
            dates[col] = d.date().isoformat()

    assert dates, "window produced no visible days"
    for i in range(len(solved)):
        row = L.CD_FIRST_ROW + i
        rank = ws.cell(row=row, column=L.CD_RANK).value
        expected = by_rank[rank]
        for col, iso in dates.items():
            got = ws.cell(row=row, column=col).value
            assert abs(got - expected.daily.get(iso, 0.0)) < TOL, f"rank {rank} {iso}"


def test_deep_dive_sums_to_high_level(values, solved):
    """Day-level and week-level views must agree for weeks inside the window."""
    window = [demo.START_WEEK + i for i in range(4)]
    for s in solved:
        for week in window:
            from gantt import calendar_utils as C
            days = [d.isoformat() for d in C.workdays(demo.YEAR, week)]
            daily = sum(s.daily.get(d, 0.0) for d in days)
            assert abs(daily - s.weekly[week]) < TOL, f"{s.parent}.{s.index} {week}"


def test_weeks_past_the_horizon_are_switched_off(values):
    """Columns exist out to WEEK_COLS but must be inert beyond the horizon."""
    ws = values[L.CALC_WEEK]
    assert L.WEEK_COLS > demo.HORIZON, "no inactive columns to test"

    for i in range(L.WEEK_COLS):
        col = L.CW_FIRST_WEEK_COL + i
        active = ws.cell(row=L.CW_ACTIVE_ROW, column=col).value
        assert active == (1 if i < demo.HORIZON else 0), f"week index {i}"
        if i >= demo.HORIZON:
            for r in range(L.CW_FIRST_ROW, L.CW_FIRST_ROW + 40):
                assert ws.cell(row=r, column=col).value == 0


def _recalc(wb, tmp_path_factory, label: str):
    src = tmp_path_factory.mktemp(label) / f"{label}.xlsx"
    wb.save(src)
    outdir = tmp_path_factory.mktemp(f"{label}_out")
    subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(src)],
        check=True, capture_output=True, timeout=300)
    return load_workbook(outdir / src.name, data_only=True)


def test_horizon_cell_extends_the_plan(built, tmp_path_factory):
    """Raising Horizon on Config must lengthen the plan with no regeneration.

    This is the whole point of building more week columns than are active: the
    12->16 case has to work by typing in a cell.
    """
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed")

    horizon = 16
    wb = load_workbook(built)
    wb[L.CONFIG].cell(row=L.CFG_HORIZON_ROW, column=2).value = horizon
    got = _recalc(wb, tmp_path_factory, "horizon16")

    ws = got[L.CALC_WEEK]
    for i in range(L.WEEK_COLS):
        col = L.CW_FIRST_WEEK_COL + i
        assert ws.cell(row=L.CW_ACTIVE_ROW, column=col).value == (1 if i < horizon else 0)

    expected = reference.schedule_weekly(reference.build_subtasks(), horizon)
    by_rank = {s.rank: s for s in expected}
    weeks = {demo.START_WEEK + i: L.CW_FIRST_WEEK_COL + i for i in range(horizon)}
    for i in range(len(expected)):
        row = L.CW_FIRST_ROW + i
        sub = by_rank[ws.cell(row=row, column=L.CW_RANK).value]
        for week, col in weeks.items():
            assert abs(ws.cell(row=row, column=col).value - sub.weekly[week]) < TOL, \
                f"rank {sub.rank} {week}"

    # the four extra weeks must actually absorb work that did not fit at 12
    tail = sum(sub.weekly[demo.START_WEEK + i]
               for sub in expected for i in range(demo.HORIZON, horizon))
    assert tail > 0, "extending the horizon scheduled nothing new"

    unfit_12 = _unfit(reference.schedule_weekly(reference.build_subtasks(), demo.HORIZON))
    unfit_16 = _unfit(expected)
    assert unfit_16 < unfit_12, "extending the horizon did not reduce unscheduled work"


def _unfit(rows) -> float:
    return sum(r.effort - sum(r.weekly.values()) for r in rows)


def test_task_beyond_horizon_is_flagged_not_silently_dropped(built, tmp_path_factory):
    """A start week past the horizon must say so, not just vanish."""
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed")

    wb = load_workbook(built)
    wb[L.TASKS].cell(row=L.TASK_FIRST_ROW + 4, column=L.T_START_WW).value = (
        demo.START_WEEK + demo.HORIZON + 2)
    got = _recalc(wb, tmp_path_factory, "beyond")

    check = got[L.TASKS].cell(row=L.TASK_FIRST_ROW + 4, column=L.T_CHECK).value
    assert "outside horizon" in check, check


def test_week_labels_wrap_at_the_year_boundary(built, tmp_path_factory):
    """WW33 + 26 weeks runs into 2027; column 21 must read WW01 '27, not WW53."""
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed")

    wb = load_workbook(built)
    wb[L.CONFIG].cell(row=L.CFG_HORIZON_ROW, column=2).value = L.WEEK_COLS
    got = _recalc(wb, tmp_path_factory, "wrap")
    ws = got[L.CALC_WEEK]

    crossed = False
    for i in range(L.WEEK_COLS):
        col = L.CW_FIRST_WEEK_COL + i
        sunday = ws.cell(row=L.CW_SUN_ROW, column=col).value.date()
        year, week = C.calendar_week(sunday)

        assert ws.cell(row=L.CW_YEAR_ROW, column=col).value == year
        assert ws.cell(row=L.CW_WEEK_ROW, column=col).value == week
        assert week <= 53, f"column {i} produced week {week}"

        expected = f"WW{week:02d}" + (f" '{year - 2000:02d}" if year != demo.YEAR else "")
        assert ws.cell(row=L.CW_LABEL_ROW, column=col).value == expected
        crossed |= year != demo.YEAR

    assert crossed, "this dataset was supposed to cross into the next year"


def test_start_week_is_a_real_calendar_week(built, tmp_path_factory):
    """Typing 2 for a plan reaching into 2027 must mean WW02 of 2027."""
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed")

    wb = load_workbook(built)
    wb[L.CONFIG].cell(row=L.CFG_HORIZON_ROW, column=2).value = L.WEEK_COLS
    # T-03 is one small sub-task, so it fits inside the weeks that remain.
    wb[L.TASKS].cell(row=L.TASK_FIRST_ROW + 2, column=L.T_START_WW).value = 2
    got = _recalc(wb, tmp_path_factory, "calweek")

    cw = got[L.CALC_WEEK]
    target = None
    for i in range(L.WEEK_COLS):
        col = L.CW_FIRST_WEEK_COL + i
        if (cw.cell(row=L.CW_WEEK_ROW, column=col).value == 2
                and cw.cell(row=L.CW_YEAR_ROW, column=col).value == demo.YEAR + 1):
            target = i + 1
    assert target is not None

    tasks = got[L.TASKS]
    assert tasks.cell(row=L.TASK_FIRST_ROW + 2, column=L.T_CHECK).value == "ok"
    assert tasks.cell(row=L.TASK_FIRST_ROW + 2, column=L.T_CALC_START).value == "WW02 '27"

    # nothing for that task before the resolved position
    row = L.CW_TASKWEEK_FIRST + 2
    for i in range(target - 1):
        assert cw.cell(row=row, column=L.CW_FIRST_WEEK_COL + i).value == 0
    assert cw.cell(row=row, column=L.CW_FIRST_WEEK_COL + target - 1).value > 0


def test_window_opened_mid_project(built, tmp_path_factory):
    """A window that starts after work has begun must show residual effort only.

    This exercises the carry-in seeding: remaining = effort minus everything
    already burned in weeks before the window.
    """
    if not Path(SOFFICE).exists():
        pytest.skip("LibreOffice not installed")

    start, weeks = demo.START_WEEK + 4, 3
    edited = tmp_path_factory.mktemp("window") / "shifted.xlsx"
    wb = load_workbook(built)
    wb[L.GANTT_DEEP][L.GD_WINDOW_START_CELL] = start
    wb[L.GANTT_DEEP][L.GD_WINDOW_WEEKS_CELL] = weeks
    wb.save(edited)

    outdir = tmp_path_factory.mktemp("window_recalc")
    subprocess.run(
        [SOFFICE, "--headless", "--norestore", "--convert-to", "xlsx",
         "--outdir", str(outdir), str(edited)],
        check=True, capture_output=True, timeout=300)
    got = load_workbook(outdir / edited.name, data_only=True)

    expected = reference.solve(window_start=start, window_weeks=weeks)
    by_rank = {s.rank: s for s in expected}

    ws = got[L.CALC_DAY]
    dates = {}
    for j in range(L.DEEP_DAY_COLS):
        col = L.CD_FIRST_DAY_COL + j
        if ws.cell(row=L.CD_INWINDOW_ROW, column=col).value:
            dates[col] = ws.cell(row=L.CD_DATE_ROW, column=col).value.date().isoformat()

    assert len(dates) == weeks * L.WORKDAYS_PER_WEEK
    for i in range(len(expected)):
        row = L.CD_FIRST_ROW + i
        rank = ws.cell(row=row, column=L.CD_RANK).value
        sub = by_rank[rank]
        for col, iso in dates.items():
            assert abs(ws.cell(row=row, column=col).value - sub.daily.get(iso, 0.0)) < TOL, \
                f"rank {rank} {iso}"

    # and the residual seeding itself
    for i in range(len(expected)):
        row = L.CD_FIRST_ROW + i
        sub = by_rank[ws.cell(row=row, column=L.CD_RANK).value]
        burned = sum(v for w, v in sub.weekly.items() if w < start)
        assert abs(ws.cell(row=row, column=L.CD_REMAINING).value
                   - max(0.0, sub.effort - burned)) < TOL


def test_holidays_remove_day_capacity(values):
    ws = values[L.CALC_DAY]
    holidays = {d for d, _ in demo.HOLIDAYS}
    flagged = set()
    for j in range(L.DEEP_DAY_COLS):
        col = L.CD_FIRST_DAY_COL + j
        d = ws.cell(row=L.CD_DATE_ROW, column=col).value
        if ws.cell(row=L.CD_HOLIDAY_ROW, column=col).value == 1:
            flagged.add(d.date().isoformat())
    # every flagged day is a real holiday
    assert flagged <= holidays
