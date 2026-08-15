"""Export of a workbook's input tabs to JSON.

The demo dataset is the oracle: `gantt/demo.py` is what the generator writes in,
so a faithful export has to give it back unchanged.
"""

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from gantt import demo, exchange, layout as L  # noqa: E402
from gantt.build import build  # noqa: E402


@pytest.fixture(scope="module")
def generated(tmp_path_factory) -> Path:
    return build(tmp_path_factory.mktemp("exchange") / "plan.xlsx")


@pytest.fixture(scope="module")
def data(generated) -> dict:
    return exchange.export(generated)


def test_export_needs_no_recalculation(generated):
    """A freshly generated workbook has no cached formula results at all.

    openpyxl does not evaluate formulas, so a exporter that read a derived
    column would return None here and a stale value from an edited file. Export
    therefore reads only cells the user types into — this asserts the premise by
    confirming the derived cells really are empty of values.
    """
    ws = load_workbook(generated, data_only=True)[L.SUBTASKS]
    derived = ws.cell(row=L.SUB_FIRST_ROW, column=L.S_EFFORT).value
    assert derived is None, "premise broken: the file now carries cached values"

    exported = exchange.export(generated)
    assert exported["sub_tasks"][0]["id"] == "T-01.01"


def test_config_round_trips(data):
    config = data["config"]
    assert config["year"] == demo.YEAR
    assert config["start_week"] == demo.START_WEEK
    assert config["horizon_weeks"] == demo.HORIZON
    assert [(c["name"], c["base_days"]) for c in config["complexity"]] == \
        [(n, d) for n, d in demo.COMPLEXITY]
    assert [(p["name"], p["order"]) for p in config["priorities"]] == \
        [(n, o) for n, o in demo.PRIORITIES]


def test_assignees_round_trip(data):
    assert [(a["name"], a["proficiency"]) for a in data["assignees"]] == \
        [(n, p) for n, p in demo.ASSIGNEES]


def test_week_grids_round_trip(data):
    """Capacity and equipment come back as the same series that went in."""
    for row in data["capacity"]:
        expected = demo.CAPACITY[row["assignee"]][:L.WEEK_COLS]
        assert row["days_by_week_offset"] == expected, row["assignee"]
    for row in data["equipment"]:
        expected = demo.EQUIPMENT[row["type"]][:L.WEEK_COLS]
        assert row["units_by_week_offset"] == expected, row["type"]


def test_capacity_rows_follow_the_assignee_order(data):
    """The Capacity name column is a formula mirroring Assignees, so it cannot
    be read; row order is the only thing tying the two together."""
    assert [c["assignee"] for c in data["capacity"]] == \
        [a["name"] for a in data["assignees"]]


def test_holidays_round_trip(data):
    assert [(h["date"], h["name"]) for h in data["holidays"]] == list(demo.HOLIDAYS)


def test_tasks_round_trip(data):
    assert len(data["tasks"]) == len(demo.TASKS)
    for got, expected in zip(data["tasks"], demo.TASKS):
        tid, name, category, priority, complexity, equipment, assignee, start, _ = expected
        assert (got["id"], got["name"], got["category"], got["priority"],
                got["complexity"], got["equipment"], got["default_assignee"],
                got["earliest_start_week"]) == \
            (tid, name, category, priority, complexity, equipment, assignee, start)


def test_sub_tasks_round_trip(data):
    expected = demo.subtasks()
    assert len(data["sub_tasks"]) == len(expected)
    for got, (parent, name, complexity, override) in zip(data["sub_tasks"], expected):
        assert got["parent"] == parent
        assert got["name"] == name
        assert got["complexity"] == complexity
        assert got["assignee"] == (override or None)


def test_sub_task_ids_match_the_workbook_rule(data):
    """Recomputed here, so they must agree with what the formula would build."""
    counts: dict[str, int] = {}
    for row in data["sub_tasks"]:
        counts[row["parent"]] = counts.get(row["parent"], 0) + 1
        assert row["id"] == f"{row['parent']}.{counts[row['parent']]:02d}"


def test_no_derived_column_leaks_into_the_export(data):
    """Derived values would be stale the moment anything upstream changed."""
    for row in data["sub_tasks"]:
        assert set(row) == {"id", "parent", "name", "complexity", "assignee"}
    for row in data["tasks"]:
        assert "effort" not in row and "check" not in row and "start_ww" not in row


def test_output_is_json_serialisable(data):
    """Dates in particular have to survive; openpyxl hands them back as datetimes."""
    text = exchange.to_json(data)
    assert json.loads(text) == data


def test_blank_rows_are_skipped(data):
    """The template pre-builds 30 task and 600 sub-task rows."""
    assert len(data["tasks"]) == 5
    assert len(data["sub_tasks"]) == 33
    assert all(row["parent"] for row in data["sub_tasks"])


def test_cli_writes_a_file(generated, tmp_path):
    out = tmp_path / "plan.json"
    assert exchange._main(["export", str(generated), "-o", str(out)]) == 0
    assert json.loads(out.read_text(encoding="utf-8"))["schema_version"] == 1


# --- Import ----------------------------------------------------------------

def _load(path):
    return load_workbook(path, data_only=False)


def test_rebuild_reproduces_the_data(generated, tmp_path):
    """export -> rebuild -> export must be a fixed point."""
    original = exchange.export(generated)
    out = tmp_path / "rebuilt.xlsx"
    exchange.rebuild(original, out)

    again = exchange.export(out)
    original.pop("source"), again.pop("source")
    assert again == original


def test_rebuild_reproduces_the_behaviour(generated, tmp_path):
    """The rebuilt workbook must *compute* the same, not merely hold the same.

    Value equality is not enough. Export renders dates as ISO strings, and
    writing one back verbatim leaves text in the cell — Excel then compares text
    against a date serial, never matches, and silently ignores every holiday.
    Nothing errors and the export still round-trips; only recalculation shows it.
    """
    from gantt import recalc

    backend = recalc.get_backend()
    if backend is None:
        pytest.skip(recalc.describe())

    out = tmp_path / "rebuilt.xlsx"
    exchange.rebuild(exchange.export(generated), out)

    before = load_workbook(backend.recalculate(generated, tmp_path / "a"),
                           data_only=True)[L.CALC_WEEK]
    after = load_workbook(backend.recalculate(out, tmp_path / "b"),
                          data_only=True)[L.CALC_WEEK]

    columns = range(L.CW_FIRST_WEEK_COL, L.CW_FIRST_WEEK_COL + L.WEEK_COLS)
    workdays_before = [before.cell(row=L.CW_WORKDAYS_ROW, column=c).value for c in columns]
    workdays_after = [after.cell(row=L.CW_WORKDAYS_ROW, column=c).value for c in columns]

    assert min(workdays_before) < L.WORKDAYS_PER_WEEK, "demo data must contain a holiday"
    assert workdays_after == workdays_before, "holidays stopped being recognised"


def test_merge_keeps_planning_decisions(generated, tmp_path):
    """The workbook owns priority; an update must not revert the planner."""
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    data["tasks"][0]["priority"] = "P3"        # as if the source system said so
    data["tasks"][0]["name"] = "Renamed upstream"

    exchange.import_data(target, data, mode="merge")
    after = exchange.export(target)

    assert after["tasks"][0]["priority"] == "P1", "merge overwrote a planning decision"
    assert after["tasks"][0]["name"] == "Renamed upstream", "merge ignored an observed fact"


def test_replace_overwrites_everything(generated, tmp_path):
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    data["tasks"][0]["priority"] = "P3"
    exchange.import_data(target, data, mode="replace")

    assert exchange.export(target)["tasks"][0]["priority"] == "P3"


def test_new_work_is_appended(generated, tmp_path):
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    data["tasks"].append({
        "id": "T-99", "name": "Discovered late", "category": "Ops",
        "priority": "P2", "complexity": "Medium", "equipment": "Rig-A",
        "default_assignee": "Bob", "earliest_start_week": 35})
    data["sub_tasks"].append({
        "parent": "T-99", "name": "Investigate", "complexity": "Simple",
        "assignee": None})

    exchange.import_data(target, data, mode="merge")
    after = exchange.export(target)

    assert [t["id"] for t in after["tasks"]][-1] == "T-99"
    added = [s for s in after["sub_tasks"] if s["parent"] == "T-99"]
    assert len(added) == 1 and added[0]["complexity"] == "Simple", \
        "a brand new row must be written in full, ownership notwithstanding"


def test_sub_tasks_match_on_parent_and_name(generated, tmp_path):
    """Positional IDs shift when a row is inserted; the pair does not.

    Matching on the ID would attach this update to whichever row happened to
    land in that position.
    """
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    data["sub_tasks"].insert(0, {"parent": "T-01", "name": "Inserted first",
                                 "complexity": "Simple", "assignee": None})
    exchange.import_data(target, data, mode="merge")

    after = exchange.export(target)
    names = [s["name"] for s in after["sub_tasks"] if s["parent"] == "T-01"]
    assert "Inserted first" in names
    assert "Collect requirements" in names, "an existing sub-task was displaced"
    assert len(names) == 4


def test_a_new_assignee_is_added_even_though_merge_owns_them(generated, tmp_path):
    """Referential integrity beats ownership: a task naming an unknown assignee
    would fail its dropdown and schedule nothing."""
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    data["assignees"].append({"name": "Dana", "proficiency": 1.1, "notes": None})
    exchange.import_data(target, data, mode="merge")

    assert "Dana" in [a["name"] for a in exchange.export(target)["assignees"]]


def test_dry_run_writes_nothing(generated, tmp_path):
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)
    before = target.read_bytes()

    data = exchange.export(target)
    data["tasks"][0]["name"] = "Changed"
    changes = exchange.import_data(target, data, mode="merge", dry_run=True)

    assert changes, "the dry run should have found a change to report"
    assert target.read_bytes() == before, "dry run modified the workbook"


def test_a_newer_schema_is_refused(generated, tmp_path):
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)
    data = exchange.export(target)
    data["schema_version"] = exchange.SCHEMA_VERSION + 1

    with pytest.raises(ValueError, match="schema"):
        exchange.import_data(target, data, mode="merge", dry_run=True)


def test_running_out_of_rows_is_reported(generated, tmp_path):
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    data = exchange.export(target)
    for i in range(L.MAX_TASKS + 5):
        data["tasks"].append({"id": f"X-{i:03d}", "name": f"Filler {i}"})

    with pytest.raises(ValueError, match="full"):
        exchange.import_data(target, data, mode="merge", dry_run=True)


def test_import_preserves_formatting(generated, tmp_path):
    """In-place import must not strip what makes the workbook readable."""
    target = tmp_path / "plan.xlsx"
    _load(generated).save(target)

    def survey(path):
        wb = load_workbook(path)
        return (len(wb.defined_names),
                sum(len(ws.conditional_formatting._cf_rules) for ws in wb.worksheets),
                sum(len(ws.data_validations.dataValidation) for ws in wb.worksheets))

    before = survey(target)
    data = exchange.export(target)
    data["tasks"][0]["name"] = "Changed"
    exchange.import_data(target, data, mode="merge")

    assert survey(target) == before


def test_replace_removes_what_the_json_does_not_mention(generated, tmp_path):
    """Rebuilding onto a fresh template must not leave the demo data behind.

    The template ships populated, so a rebuild from real data lands on top of
    Alice, Bob, Carol, Rig-A and Rig-B. Anything the JSON omits has to be
    cleared, or a migration silently mixes demo rows into live data — and a
    stale capacity figure under a cleared name is invisible but still counted.
    """
    data = exchange.export(generated)
    data["assignees"] = data["assignees"][:1]
    data["capacity"] = data["capacity"][:1]
    data["equipment"] = data["equipment"][:1]
    data["tasks"] = data["tasks"][:2]
    keep = {t["id"] for t in data["tasks"]}
    data["sub_tasks"] = [s for s in data["sub_tasks"] if s["parent"] in keep]

    out = tmp_path / "smaller.xlsx"
    exchange.rebuild(data, out)
    after = exchange.export(out)

    assert [a["name"] for a in after["assignees"]] == ["Alice"]
    assert [e["type"] for e in after["equipment"]] == ["Rig-A"]
    assert [t["id"] for t in after["tasks"]] == ["T-01", "T-02"]
    assert {s["parent"] for s in after["sub_tasks"]} == keep

    ws = load_workbook(out)[L.CAPACITY]
    for row in range(L.GRID_FIRST_DATA_ROW + 1, L.GRID_FIRST_DATA_ROW + 3):
        values = [ws.cell(row=row, column=c).value
                  for c in range(L.GRID_FIRST_WEEK_COL,
                                 L.GRID_FIRST_WEEK_COL + L.WEEK_COLS)]
        assert all(v is None for v in values), \
            f"stale capacity left on row {row} under a cleared assignee"


def test_merge_never_removes_anything(generated, tmp_path):
    """The mirror image: a weekly update that omits a task must not delete it."""
    target = tmp_path / "plan.xlsx"
    load_workbook(generated).save(target)

    data = exchange.export(target)
    data["tasks"] = data["tasks"][:1]
    data["sub_tasks"] = data["sub_tasks"][:1]
    exchange.import_data(target, data, mode="merge")

    after = exchange.export(target)
    assert len(after["tasks"]) == 5, "merge deleted tasks the update did not mention"
    assert len(after["sub_tasks"]) == 33
