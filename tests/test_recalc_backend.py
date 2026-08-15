"""Checks on the recalculation backend itself.

These matter because the numeric half of the suite is only as trustworthy as the
engine behind it, and the engine differs by platform. A test that passes against
LibreOffice while the file is opened in Excel has proved less than it appears to.
"""

import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from gantt import recalc  # noqa: E402


def test_selection_honours_the_override(monkeypatch):
    monkeypatch.setenv(recalc.ENV_BACKEND, "none")
    assert recalc.get_backend() is None


def test_excel_is_never_selected_off_windows(monkeypatch):
    monkeypatch.setattr(recalc.platform, "system", lambda: "Darwin")
    assert recalc.excel_available() is False


def test_windows_looks_in_windows_locations(monkeypatch):
    """The search must not depend on the platform it was written on."""
    monkeypatch.delenv(recalc.ENV_SOFFICE, raising=False)
    monkeypatch.setattr(recalc.platform, "system", lambda: "Windows")
    monkeypatch.setattr(recalc.shutil, "which", lambda _name: None)

    seen = []

    class FakePath(type(Path())):
        def exists(self):
            seen.append(str(self))
            return False

    monkeypatch.setattr(recalc, "Path", FakePath)
    assert recalc.find_libreoffice() is None
    assert any(s.lower().endswith("soffice.exe") for s in seen), seen


def test_a_bad_explicit_path_is_not_silently_ignored(monkeypatch):
    monkeypatch.setenv(recalc.ENV_SOFFICE, "/definitely/not/here/soffice")
    assert recalc.find_libreoffice() is None


def test_the_engine_that_ran_is_the_one_that_was_selected(tmp_path):
    """Proves the backend really recalculated, and which one did.

    Every engine stamps its name into docProps/app.xml, so a computed value
    shows something evaluated the formula and the stamp shows what. Without
    this, `GANTT_RECALC=excel` silently falling back would go unnoticed and the
    numeric tests would quietly be verifying the wrong engine.
    """
    backend = recalc.get_backend()
    if backend is None:
        pytest.skip(recalc.describe())

    report = recalc.verify(tmp_path)
    assert report["computed"] == 2, f"nothing recalculated: {report}"

    engine = report["fingerprint"].lower()
    expected = {"excel": "microsoft excel", "libreoffice": "libreoffice"}[backend.name]
    assert expected in engine, (
        f"selected {backend.name} but the file was written by "
        f"{report['fingerprint']!r}")


def test_conditional_formatting_survives_a_foreign_save(tmp_path):
    """A guard for the in-place import planned in Phase 2 of the roadmap.

    Import writes values into the user's own workbook with openpyxl, and by then
    that workbook will have been opened and saved by Excel. Excel rewrites some
    rule types into an x14 extension block that openpyxl does not understand and
    warns about, and anything it does not understand it drops on save. If that
    happens, editing a user's file in place would silently strip its colours and
    the import strategy needs rethinking.

    LibreOffice emits no extensions for these rules, so on macOS this passes
    trivially. It only carries information when run on Windows against Excel.
    """
    from openpyxl import load_workbook

    from gantt.build import build

    backend = recalc.get_backend()
    if backend is None:
        pytest.skip(recalc.describe())

    generated = build(tmp_path / "generated.xlsx")
    before = _cf_ranges(generated)

    foreign = backend.recalculate(generated, tmp_path / "foreign")
    after_engine = _cf_ranges(foreign)

    resaved = tmp_path / "resaved.xlsx"
    load_workbook(foreign).save(resaved)
    after_openpyxl = _cf_ranges(resaved)

    assert after_engine == before, (
        f"{backend.name} changed the rule count on save: {before} -> {after_engine}")
    assert after_openpyxl == before, (
        f"openpyxl dropped rules re-saving a file {backend.name} had written: "
        f"{before} -> {after_openpyxl}. In-place import would strip formatting.")


def _cf_ranges(path) -> int:
    from openpyxl import load_workbook
    book = load_workbook(path)
    return sum(len(ws.conditional_formatting._cf_rules) for ws in book.worksheets)
